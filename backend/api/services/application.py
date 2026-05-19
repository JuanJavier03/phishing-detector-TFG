from __future__ import annotations

"""
Servicio principal de aplicacion.

Coordina uploads, persistencia, jobs de enriquecimiento, recomputo MCDM,
serializacion para la API y exportaciones analiticas.
"""

from concurrent.futures import ThreadPoolExecutor, as_completed
from copy import deepcopy
from io import BytesIO
import re
import statistics
import time
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from uuid import UUID

from fastapi import BackgroundTasks, UploadFile
from sqlalchemy import Select, or_, select
from sqlalchemy.orm import Session, load_only, selectinload

from api.db.config import get_analysis_batch_workers
from api.db.models import ApiCallMetric, Batch, Email, EmailAnalysisMetric, EmailEnrichment, EmailMcdmVector, Job, McdmRuntimeMetric
from api.db.session import get_session_local
from api.services.chart_value_extractors import (
    build_label_bins,
    build_label_segments,
    build_numeric_points,
    build_score_band_bins,
    build_semantic_histogram_bins,
    chart_config_for_subcriterion,
    extract_chart_display_record,
)
from api.services.subcriteria_catalog import (
    SUBCRITERIA,
    available_subcriteria,
    get_subcriterion_definition,
)
from api.services.upload_pipeline import parse_uploaded_email
from scripts.parse_emails import sha256_bytes
from scripts.enrichment.external_api_errors import FatalExternalApiError
from scripts.enrichment.api_call_tracking import collect_api_calls
from scripts.enrichment.enrich_all import enrich_email_in_data, normalize_selected_subcriteria
from scripts.enrichment.mcdm_score_helper import (
    compute_mcdm_block_for_email_with_hardcoded_references,
    compute_mcdm_blocks_for_emails,
    field_unit_score_with_context,
)
from scripts.enrichment.vector_schema import MCDM_SUBCRITERIA_WEIGHTS, SUBCRITERIA_DEFINITIONS, VECTOR_FIELD_ORDER
from scripts.enrichment.link_domain_utils import (
    LINK_API_SKIP_REASON,
    LINK_API_SKIP_THRESHOLD,
    MAX_LINK_API_LOOKUPS_PER_EMAIL,
)
from scripts.utils.auth_header_scoring import (
    DKIM_SCORE_MAP,
    DMARC_AUTH_ALIGNMENT_RULE_VERSION,
    DMARC_FAIL_POLICY_SCORE_MAP,
    DMARC_PASS_WITH_AUTH_FAILURE_SCORE,
    DMARC_STATUS_SCORE_MAP,
    SPF_SCORE_MAP,
)
from scripts.utils.subcriteria_utils import normalize_subcriterion_result


ENRICHMENT_JSON_COLUMNS = {item["key"]: item["enrichment_column"] for item in SUBCRITERIA}


VECTOR_FIELD_LABELS = {item["vector_field"]: item["label"] for item in SUBCRITERIA}
ALL_VECTOR_FIELDS = list(VECTOR_FIELD_ORDER)

EXPORT_BASE_HEADERS = [
    "email_id",
    "email_name",
    "subject",
    "status",
    "mcdm_score",
]

MALICIOUS_DATASET_EMAIL_NAME_PREFIXES = ("sample", "nazario")
BENIGN_DATASET_EMAIL_NAME_PREFIXES = ("benigno",)


CRITERION_IDENTIFIERS = {
    item["key"]: f'{item["family"]}.{item["key"]}'
    for item in SUBCRITERIA
}


AGE_SUBCRITERIA_KEYS = {
    "domain_age",
    "routing_domain_age",
    "link_domain_age",
}


SUPPORTED_BATCH_CHART_TYPES = {
    "histogram",
    "donut",
    "pie",
    "bubble_lane",
    "pyramid",
    "stacked_bar",
    "box_plot",
    "waffle",
    "band_bars",
}


UPLOAD_READ_CHUNK_BYTES = 1024 * 1024
EMAIL_PROCESSING_STATE_READY = "ready"
EMAIL_PROCESSING_STATE_ERROR = "error"
EMAIL_PROCESSING_STATE_CANCELLED = "cancelled"
EMAIL_DISCARDED_STATES = {
    EMAIL_PROCESSING_STATE_ERROR,
    EMAIL_PROCESSING_STATE_CANCELLED,
}


JOB_LOAD_ONLY_COLUMNS = (
    Job.id,
    Job.job_type,
    Job.target_type,
    Job.target_id,
    Job.email_id,
    Job.batch_id,
    Job.selected_subcriteria,
    Job.status,
    Job.progress_current,
    Job.progress_total,
    Job.error_message,
    Job.created_at,
    Job.started_at,
    Job.finished_at,
)


SEMANTIC_VALUE_LABELS_BY_SUBCRITERION_KEY: Dict[str, Dict[float, str]] = {
    "domain_vs_ip_country": {
        0.0: "Match",
        0.5: "Sin datos",
        1.0: "Mismatch",
    },
    "ip_reputation": {
        0.0: "No listada",
        1.0: "Listada",
    },
    "domain_reputation": {
        0.0: "No malicioso",
        1.0: "Malicioso",
    },
    "from_return_path_subdomain_match": {
        0.0: "Coincide",
        1.0: "Mismatch",
    },
    "routing_country_mismatch": {0.0: "No", 1.0: "Si"},
    "routing_domain_reputation": {0.0: "No malicioso", 1.0: "Malicioso"},
    "routing_ip_reputation": {0.0: "No listada", 1.0: "Listada"},
    "body_obfuscation_base64": {
        0.0: "No",
        1.0: "Si",
    },
    "link_domain_country_vs_modal": {
        0.0: "Match",
        1.0: "Mismatch",
    },
    "link_domain_reputation": {
        0.0: "No malicioso",
        1.0: "Malicioso",
    },
    "link_domain_match_modal": {
        0.0: "Coincide",
        1.0: "Mismatch",
    },
    "link_captcha": {
        0.0: "No",
        1.0: "Si",
    },
    "sender_numeric_subdomain": {
        0.0: "0 labels",
        1.0: "1 label",
        2.0: "2+ labels",
    },
    "link_numeric_subdomain": {
        0.0: "0 labels",
        1.0: "1 label",
        2.0: "2+ labels",
    },
}


def _sanitize_selected_subcriteria(values: Optional[Iterable[Any]]) -> List[str]:
    if values is None:
        return []

    cleaned: List[str] = []
    seen = set()
    for raw_value in values:
        if not isinstance(raw_value, str):
            continue
        key = raw_value.strip()
        if not key or key not in ENRICHMENT_JSON_COLUMNS or key in seen:
            continue
        seen.add(key)
        cleaned.append(key)
    return cleaned


def _merge_selected_subcriteria(existing: Optional[Iterable[Any]], incoming: Optional[Iterable[Any]]) -> List[str]:
    merged = _sanitize_selected_subcriteria(existing)
    seen = set(merged)
    for key in _sanitize_selected_subcriteria(incoming):
        if key in seen:
            continue
        seen.add(key)
        merged.append(key)
    return merged


def _effective_selected_subcriteria(email: Email) -> List[str]:
    selected = _sanitize_selected_subcriteria(email.selected_subcriteria)
    for job in email.jobs:
        selected = _merge_selected_subcriteria(selected, job.selected_subcriteria)
    return selected


def _email_is_discarded(email: Optional[Email]) -> bool:
    return bool(email is not None and email.processing_state in EMAIL_DISCARDED_STATES)


def _processable_batch_emails(batch: Batch) -> List[Email]:
    return [email for email in batch.emails if not _email_is_discarded(email)]


def _retryable_batch_emails(batch: Batch) -> List[Email]:
    return [email for email in batch.emails if _email_is_discarded(email)]


def _email_processing_error_payload(email: Email) -> Optional[Dict[str, Any]]:
    if not _email_is_discarded(email):
        return None
    return {
        "state": email.processing_state,
        "code": email.processing_error_code,
        "message": email.processing_error_message,
        "at": _utc_iso(email.processing_error_at),
    }


def _all_subcriteria_keys() -> List[str]:
    return [item["key"] for item in available_subcriteria()]


def _missing_subcriteria_for_email(email: Email) -> List[str]:
    if _email_is_discarded(email):
        return []
    selected = set(_effective_selected_subcriteria(email))
    return [key for key in _all_subcriteria_keys() if key not in selected]


def _missing_subcriteria_for_batch(batch: Batch) -> List[str]:
    selected = {
        key
        for email in _processable_batch_emails(batch)
        for key in _effective_selected_subcriteria(email)
    }
    return [key for key in _all_subcriteria_keys() if key not in selected]


def _utc_iso(value: Optional[datetime]) -> Optional[str]:
    return value.isoformat() if isinstance(value, datetime) else None


def _clear_email_enrichment_outputs(email: Email) -> None:
    if email.enrichment is None:
        return

    for column in ENRICHMENT_JSON_COLUMNS.values():
        setattr(email.enrichment, column, None)
    email.enrichment.mcdm_score = None
    email.enrichment.mcdm_is_mock = True
    email.enrichment.mcdm_method = None
    email.enrichment.updated_at = datetime.utcnow()

    vector = email.enrichment.vector
    if vector is None:
        return
    for field in ALL_VECTOR_FIELDS:
        setattr(vector, field, None)
    vector.updated_at = datetime.utcnow()


def _reset_email_processing_state(email: Email) -> None:
    email.processing_state = EMAIL_PROCESSING_STATE_READY
    email.processing_error_code = None
    email.processing_error_message = None
    email.processing_error_at = None

    raw_json = dict(email.raw_json or {})
    raw_json.pop("enrichment", None)
    raw_json.pop("numeric_values", None)
    raw_json.pop("mcdm", None)
    raw_json.pop("processing_error", None)
    email.raw_json = raw_json
    _clear_email_enrichment_outputs(email)


def _mark_email_as_processing_state(
    email: Email,
    *,
    processing_state: str,
    error_code: str,
    error_message: str,
) -> None:
    timestamp = datetime.utcnow()
    email.processing_state = processing_state
    email.processing_error_code = error_code[:128]
    email.processing_error_message = error_message
    email.processing_error_at = timestamp

    raw_json = dict(email.raw_json or {})
    raw_json.pop("enrichment", None)
    raw_json.pop("numeric_values", None)
    raw_json.pop("mcdm", None)
    raw_json["processing_error"] = {
        "state": processing_state,
        "code": error_code[:128],
        "message": error_message,
        "at": _utc_iso(timestamp),
    }
    email.raw_json = raw_json
    _clear_email_enrichment_outputs(email)


def _mark_email_as_processing_error(
    email: Email,
    *,
    error_code: str,
    error_message: str,
) -> None:
    _mark_email_as_processing_state(
        email,
        processing_state=EMAIL_PROCESSING_STATE_ERROR,
        error_code=error_code,
        error_message=error_message,
    )


def _mark_email_as_processing_cancelled(
    email: Email,
    *,
    error_code: str,
    error_message: str,
) -> None:
    _mark_email_as_processing_state(
        email,
        processing_state=EMAIL_PROCESSING_STATE_CANCELLED,
        error_code=error_code,
        error_message=error_message,
    )


def _safe_filename(value: str) -> str:
    sanitized = re.sub(r'[<>:"/\\|?*]+', "_", value).strip()
    return sanitized or "batch"


def _datetime_sort_value(value: Optional[datetime]) -> float:
    if not isinstance(value, datetime):
        return 0.0
    try:
        return value.timestamp()
    except Exception:
        return 0.0


def _email_merge_identity(email: Email) -> str:
    raw_json = email.raw_json if isinstance(email.raw_json, dict) else {}
    metadata = raw_json.get("metadata") if isinstance(raw_json.get("metadata"), dict) else {}
    raw_sha = metadata.get("raw_sha256")
    if isinstance(raw_sha, str) and raw_sha.strip():
        return f"raw_sha256:{raw_sha.strip().lower()}"

    headers = raw_json.get("headers") if isinstance(raw_json.get("headers"), dict) else {}
    message_id = headers.get("message_id")
    if isinstance(message_id, str) and message_id.strip():
        return f"message_id:{message_id.strip().lower()}"

    return f"email_id:{email.id}"


def _email_evaluation_sort_key(email: Email) -> Tuple[float, str]:
    candidates: List[datetime] = []
    if email.enrichment is not None:
        candidates.append(email.enrichment.updated_at)
    for job in email.jobs:
        candidates.extend([job.finished_at, job.started_at, job.created_at])
    candidates.append(email.created_at)
    latest = max(candidates, key=_datetime_sort_value, default=None)
    return _datetime_sort_value(latest), str(email.id)


def _normalize_vector_field_value(field: str, value: Any) -> Any:
    if field == "c2_obfuscation_base64_present" and isinstance(value, (int, float)):
        return 1 if value > 0 else 0
    return value


def _model_to_vector_dict(vector: Optional[EmailMcdmVector]) -> Dict[str, Any]:
    if vector is None:
        return {}
    data: Dict[str, Any] = {}
    for field in VECTOR_FIELD_LABELS:
        data[field] = _normalize_vector_field_value(field, getattr(vector, field))
    return data


def _selected_complete(email: Email) -> Tuple[bool, bool]:
    if _email_is_discarded(email):
        return False, False
    selected = _effective_selected_subcriteria(email)
    if not selected:
        return False, False
    enrichment = email.enrichment
    if enrichment is None:
        return False, False
    present = 0
    for key in selected:
        column = ENRICHMENT_JSON_COLUMNS[key]
        if getattr(enrichment, column) is not None:
            present += 1
    return present == len(selected), present > 0


def _normalize_stored_subcriterion_result(key: str, result: Any) -> Optional[Dict[str, Any]]:
    if not isinstance(result, dict):
        return None
    criterion = CRITERION_IDENTIFIERS.get(key)
    if not criterion:
        return result
    normalized = normalize_subcriterion_result(result, criterion=criterion)
    return normalized if isinstance(normalized, dict) else result


def _subcriterion_has_insufficient_data(result: Optional[Dict[str, Any]]) -> bool:
    if not isinstance(result, dict):
        return False
    detail = result.get("detail")
    if not isinstance(detail, dict):
        return False
    if detail.get("insufficient_data") is True:
        return True
    meta = detail.get("meta")
    if isinstance(meta, dict) and meta.get("insufficient_data") is True:
        return True
    return False


def _display_subcriterion_value(key: str, result: Optional[Dict[str, Any]]) -> Any:
    if not isinstance(result, dict):
        return None
    value = result.get("value")
    if key in AGE_SUBCRITERIA_KEYS and value == 0 and _subcriterion_has_insufficient_data(result):
        return None
    return value


def _subcriterion_status(
    email: Email,
    key: str,
    result: Optional[Dict[str, Any]],
    *,
    selected: Optional[set[str]] = None,
) -> str:
    if isinstance(result, dict):
        return "completed"

    selected_keys = selected if selected is not None else set(_effective_selected_subcriteria(email))
    if key not in selected_keys:
        return "available"

    if email.processing_state == EMAIL_PROCESSING_STATE_CANCELLED:
        return "cancelled"
    if email.processing_state == EMAIL_PROCESSING_STATE_ERROR:
        return "error"
    return "not_analyzed"


def _latest_job_for_email(db: Session, email: Email) -> Optional[Job]:
    loaded_jobs: List[Job] = list(email.jobs) if "jobs" in email.__dict__ else []
    if email.batch is not None and "jobs" in email.batch.__dict__:
        loaded_jobs.extend(email.batch.jobs)
    if loaded_jobs:
        return max(
            loaded_jobs,
            key=lambda job: (
                job.created_at or datetime.min,
                str(job.id),
            ),
        )

    if email.batch_id:
        stmt = select(Job).where(or_(Job.email_id == email.id, Job.batch_id == email.batch_id)).order_by(Job.created_at.desc())
    else:
        stmt = select(Job).where(Job.email_id == email.id).order_by(Job.created_at.desc())
    return db.scalar(stmt)


def _latest_job_for_batch(db: Session, batch_id: UUID) -> Optional[Job]:
    batch = db.get(Batch, batch_id)
    if batch is not None and "jobs" in batch.__dict__ and batch.jobs:
        return max(
            batch.jobs,
            key=lambda job: (
                job.created_at or datetime.min,
                str(job.id),
            ),
        )

    stmt = select(Job).where(Job.batch_id == batch_id).order_by(Job.created_at.desc())
    return db.scalar(stmt)


def _job_is_active(job: Optional[Job]) -> bool:
    return job is not None and job.status in {"queued", "running"}


def _batch_has_active_email_jobs(batch: Batch) -> bool:
    return any(
        _job_is_active(job)
        for email in _processable_batch_emails(batch)
        for job in email.jobs
    )


def _serialize_job(job: Optional[Job]) -> Optional[Dict[str, Any]]:
    if job is None:
        return None
    return {
        "id": str(job.id),
        "job_type": job.job_type,
        "target_type": job.target_type,
        "target_id": job.target_id,
        "email_id": str(job.email_id) if job.email_id else None,
        "batch_id": str(job.batch_id) if job.batch_id else None,
        "selected_subcriteria": _sanitize_selected_subcriteria(job.selected_subcriteria),
        "status": job.status,
        "progress_current": job.progress_current,
        "progress_total": job.progress_total,
        "error_message": job.error_message,
        "created_at": _utc_iso(job.created_at),
        "started_at": _utc_iso(job.started_at),
        "finished_at": _utc_iso(job.finished_at),
    }


def _email_status(db: Session, email: Email) -> str:
    if email.processing_state == EMAIL_PROCESSING_STATE_CANCELLED:
        return "cancelled"
    if email.processing_state == EMAIL_PROCESSING_STATE_ERROR:
        return "error"
    all_done, any_done = _selected_complete(email)
    if all_done:
        return "completed"
    latest_job = _latest_job_for_email(db, email)
    if _job_is_active(latest_job):
        return latest_job.status
    if any_done:
        return "partial"
    return "pending"


def _serialize_subcriteria_summary(email: Email) -> List[Dict[str, Any]]:
    selected = _effective_selected_subcriteria(email)
    enrichment = email.enrichment
    vector_values = _model_to_vector_dict(enrichment.vector if enrichment else None)
    items: List[Dict[str, Any]] = []
    selected_keys = set(selected)
    for item in available_subcriteria():
        key = item["key"]
        result = getattr(enrichment, item["enrichment_column"]) if enrichment is not None else None
        normalized_result = _normalize_stored_subcriterion_result(key, result)
        display_value = _display_subcriterion_value(key, normalized_result)
        is_selected = key in selected_keys
        status = _subcriterion_status(email, key, normalized_result, selected=selected_keys)
        mcdm_score = None
        if isinstance(normalized_result, dict):
            mcdm_score = field_unit_score_with_context(item["vector_field"], display_value, vector_values)
        items.append(
            {
                "key": key,
                "label": item["label"],
                "family": item["family"],
                "vector_field": item["vector_field"],
                "value_type": item["value_type"],
                "mcdm_objective": item["mcdm_objective"],
                "mcdm_weight": item["mcdm_weight"],
                "measurement": item["measurement"],
                "value": display_value,
                "mcdm_score": mcdm_score,
                "updated_at": normalized_result.get("updated_at") if isinstance(normalized_result, dict) else None,
                "has_result": isinstance(normalized_result, dict),
                "selected": is_selected,
                "status": status,
            }
        )
    return items


def _append_export_sheet(
    workbook: Any,
    title: str,
    batch: Batch,
    definitions: Sequence[Dict[str, Any]],
    db: Session,
    *,
    score_mode: bool,
) -> None:
    from openpyxl.styles import Font, PatternFill

    worksheet = workbook.create_sheet(title=title)
    headers = list(EXPORT_BASE_HEADERS)
    headers.extend(item["key"] for item in definitions)
    worksheet.append(headers)
    malicious_row_fill = PatternFill(fill_type="solid", fgColor="FDE2E1")
    benign_row_fill = PatternFill(fill_type="solid", fgColor="DDEFD8")

    emails = sorted(
        _processable_batch_emails(batch),
        key=lambda email: (
            email.enrichment is None or email.enrichment.mcdm_score is None,
            -(email.enrichment.mcdm_score or 0.0),
            email.created_at or datetime.min,
            str(email.id),
        ),
    )
    for email in emails:
        vector_values = _model_to_vector_dict(email.enrichment.vector if email.enrichment else None)
        row = [
            str(email.id),
            email.name,
            email.subject,
            _email_status(db, email),
            email.enrichment.mcdm_score if email.enrichment else None,
        ]
        for definition in definitions:
            value = vector_values.get(definition["vector_field"])
            if score_mode:
                row.append(field_unit_score_with_context(definition["vector_field"], value, vector_values))
            else:
                row.append(value)
        worksheet.append(row)
        email_name = email.name.strip().lower() if isinstance(email.name, str) else ""
        row_fill = None
        if email_name.startswith(MALICIOUS_DATASET_EMAIL_NAME_PREFIXES):
            row_fill = malicious_row_fill
        elif email_name.startswith(BENIGN_DATASET_EMAIL_NAME_PREFIXES):
            row_fill = benign_row_fill
        if row_fill is not None:
            for cell in worksheet[worksheet.max_row]:
                cell.fill = row_fill

    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.font = header_font

    worksheet.freeze_panes = "A2"

    for column_cells in worksheet.columns:
        values = [cell.value for cell in column_cells if cell.value is not None]
        width = max((len(str(value)) for value in values), default=10)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 48)


def _endpoint_summary(api_calls: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    summary: Dict[Tuple[str, str, str], int] = {}
    for call in api_calls:
        key = (
            str(call.get("provider") or ""),
            str(call.get("endpoint") or ""),
            str(call.get("method") or ""),
        )
        summary[key] = summary.get(key, 0) + 1
    return [
        {
            "provider": provider,
            "endpoint": endpoint,
            "method": method,
            "count": count,
        }
        for (provider, endpoint, method), count in sorted(summary.items())
    ]


def _persist_email_analysis_metric(
    db: Session,
    *,
    job: Job,
    email: Email,
    started_at: datetime,
    duration_ms: float,
    status: str,
    selected_subcriteria: Sequence[str],
    api_calls: Sequence[Dict[str, Any]],
    error_message: Optional[str] = None,
) -> EmailAnalysisMetric:
    finished_at = datetime.utcnow()
    metric = EmailAnalysisMetric(
        job_id=job.id,
        email_id=email.id,
        batch_id=email.batch_id or job.batch_id,
        job_type=job.job_type,
        selected_subcriteria=list(selected_subcriteria),
        status=status,
        started_at=started_at,
        finished_at=finished_at,
        duration_ms=max(float(duration_ms), 0.0),
        subcriteria_count=len(selected_subcriteria),
        api_call_count=len(api_calls),
        api_endpoints=_endpoint_summary(api_calls),
        error_message=error_message,
    )
    db.add(metric)
    db.flush()
    for call in api_calls:
        db.add(
            ApiCallMetric(
                analysis_metric_id=metric.id,
                job_id=job.id,
                email_id=email.id,
                batch_id=email.batch_id or job.batch_id,
                provider=str(call.get("provider") or "unknown")[:64],
                endpoint=str(call.get("endpoint") or "unknown")[:512],
                method=str(call.get("method") or "GET")[:16],
                url=call.get("url"),
                target=call.get("target"),
                http_status=call.get("http_status"),
                success=call.get("success"),
                error=call.get("error"),
                duration_ms=max(float(call.get("duration_ms") or 0.0), 0.0),
            )
        )
    return metric


def _persist_mcdm_runtime_metric(
    db: Session,
    *,
    job: Optional[Job],
    batch_id: Optional[UUID],
    operation: str,
    started_at: datetime,
    duration_ms: float,
    email_count: int,
    status: str = "completed",
    error_message: Optional[str] = None,
) -> McdmRuntimeMetric:
    duration = max(float(duration_ms), 0.0)
    metric = McdmRuntimeMetric(
        job_id=job.id if job is not None else None,
        batch_id=batch_id,
        operation=operation,
        status=status,
        email_count=max(int(email_count), 0),
        started_at=started_at,
        finished_at=datetime.utcnow(),
        duration_ms=duration,
        duration_per_email_ms=(duration / email_count) if email_count > 0 else None,
        error_message=error_message,
    )
    db.add(metric)
    db.flush()
    return metric


def _seconds(ms: Optional[float]) -> Optional[float]:
    if ms is None:
        return None
    return round(float(ms) / 1000.0, 6)


def _stats(values: Sequence[float]) -> Dict[str, Optional[float]]:
    if not values:
        return {"total": None, "min": None, "max": None, "mean": None, "median": None}
    return {
        "total": float(sum(values)),
        "min": float(min(values)),
        "max": float(max(values)),
        "mean": float(statistics.mean(values)),
        "median": float(statistics.median(values)),
    }


def _load_export_analysis_metrics(db: Session, batch: Batch) -> List[EmailAnalysisMetric]:
    if not hasattr(db, "scalars"):
        return []
    email_ids = [email.id for email in _processable_batch_emails(batch)]
    if not email_ids:
        return []
    stmt = (
        select(EmailAnalysisMetric)
        .where(or_(EmailAnalysisMetric.batch_id == batch.id, EmailAnalysisMetric.email_id.in_(email_ids)))
        .order_by(EmailAnalysisMetric.started_at.asc(), EmailAnalysisMetric.created_at.asc())
    )
    return list(db.scalars(stmt).unique().all())


def _load_export_api_call_metrics(db: Session, metrics: Sequence[EmailAnalysisMetric]) -> List[ApiCallMetric]:
    if not hasattr(db, "scalars") or not metrics:
        return []
    metric_ids = [metric.id for metric in metrics]
    stmt = (
        select(ApiCallMetric)
        .where(ApiCallMetric.analysis_metric_id.in_(metric_ids))
        .order_by(ApiCallMetric.created_at.asc(), ApiCallMetric.id.asc())
    )
    return list(db.scalars(stmt).unique().all())


def _load_export_mcdm_runtime_metrics(db: Session, batch: Batch) -> List[McdmRuntimeMetric]:
    if not hasattr(db, "scalars"):
        return []
    stmt = (
        select(McdmRuntimeMetric)
        .where(McdmRuntimeMetric.batch_id == batch.id)
        .order_by(McdmRuntimeMetric.started_at.asc(), McdmRuntimeMetric.created_at.asc())
    )
    return list(db.scalars(stmt).unique().all())


def _style_export_worksheet(worksheet: Any) -> None:
    from openpyxl.styles import Font

    for row in worksheet.iter_rows():
        first = row[0].value if row else None
        if first in {
            "Resumen",
            "Detalle por correo",
            "Resumen por endpoint",
            "Detalle de llamadas",
            "Configuracion enlaces",
            "Pesos MCDM",
            "Rangos MCDM",
            "Mapeos SPF/DKIM/DMARC",
            "Detalle MCDM",
        }:
            for cell in row:
                cell.font = Font(bold=True)
        elif first and all(cell.value is not None for cell in row[:1]):
            pass

    for column_cells in worksheet.columns:
        values = [cell.value for cell in column_cells if cell.value is not None]
        width = max((len(str(value)) for value in values), default=10)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 64)


def _append_timing_export_sheet(workbook: Any, batch: Batch, metrics: Sequence[EmailAnalysisMetric]) -> None:
    worksheet = workbook.create_sheet(title="Tiempos")
    email_by_id = {str(email.id): email for email in _processable_batch_emails(batch)}
    durations = [float(metric.duration_ms) for metric in metrics]
    api_counts = [int(metric.api_call_count or 0) for metric in metrics]
    duration_stats = _stats(durations)
    api_stats = _stats([float(value) for value in api_counts])

    worksheet.append(["Resumen", None])
    worksheet.append(["analisis_total", len(metrics)])
    worksheet.append(["tiempo_total_s", _seconds(duration_stats["total"])])
    worksheet.append(["tiempo_min_s", _seconds(duration_stats["min"])])
    worksheet.append(["tiempo_max_s", _seconds(duration_stats["max"])])
    worksheet.append(["tiempo_media_s", _seconds(duration_stats["mean"])])
    worksheet.append(["tiempo_mediana_s", _seconds(duration_stats["median"])])
    worksheet.append(["llamadas_api_total", sum(api_counts)])
    worksheet.append(["llamadas_api_min", api_stats["min"]])
    worksheet.append(["llamadas_api_max", api_stats["max"]])
    worksheet.append(["llamadas_api_media", api_stats["mean"]])
    worksheet.append(["llamadas_api_mediana", api_stats["median"]])
    worksheet.append([])
    worksheet.append(["Detalle por correo"])
    worksheet.append(
        [
            "analysis_id",
            "job_id",
            "job_type",
            "email_id",
            "email_name",
            "subject",
            "status",
            "started_at",
            "finished_at",
            "duration_s",
            "subcriteria_count",
            "selected_subcriteria",
            "api_call_count",
            "api_endpoints",
            "error_message",
        ]
    )
    for metric in metrics:
        email = email_by_id.get(str(metric.email_id))
        worksheet.append(
            [
                str(metric.id),
                str(metric.job_id) if metric.job_id else None,
                metric.job_type,
                str(metric.email_id),
                email.name if email else None,
                email.subject if email else None,
                metric.status,
                _utc_iso(metric.started_at),
                _utc_iso(metric.finished_at),
                _seconds(metric.duration_ms),
                metric.subcriteria_count,
                ", ".join(metric.selected_subcriteria or []),
                metric.api_call_count,
                ", ".join(
                    f'{item.get("provider")}:{item.get("endpoint")} ({item.get("count")})'
                    for item in (metric.api_endpoints or [])
                    if isinstance(item, dict)
                ),
                metric.error_message,
            ]
        )
    worksheet.freeze_panes = "A16"
    _style_export_worksheet(worksheet)


def _append_mcdm_runtime_export_sheet(
    workbook: Any,
    metrics: Sequence[McdmRuntimeMetric],
) -> None:
    worksheet = workbook.create_sheet(title="Tiempos MCDM")
    durations = [float(metric.duration_ms) for metric in metrics]
    per_email_durations = [
        float(metric.duration_per_email_ms)
        for metric in metrics
        if metric.duration_per_email_ms is not None
    ]
    duration_stats = _stats(durations)
    per_email_stats = _stats(per_email_durations)

    worksheet.append(["Resumen", None])
    worksheet.append(["calculos_total", len(metrics)])
    worksheet.append(["nota", "El MCDM se calcula globalmente sobre la matriz del lote; el dato por correo es derivado: tiempo_global / numero_correos."])
    worksheet.append(["tiempo_global_total_s", _seconds(duration_stats["total"])])
    worksheet.append(["tiempo_global_min_s", _seconds(duration_stats["min"])])
    worksheet.append(["tiempo_global_max_s", _seconds(duration_stats["max"])])
    worksheet.append(["tiempo_global_media_s", _seconds(duration_stats["mean"])])
    worksheet.append(["tiempo_global_mediana_s", _seconds(duration_stats["median"])])
    worksheet.append(["tiempo_derivado_por_correo_min_s", _seconds(per_email_stats["min"])])
    worksheet.append(["tiempo_derivado_por_correo_max_s", _seconds(per_email_stats["max"])])
    worksheet.append(["tiempo_derivado_por_correo_media_s", _seconds(per_email_stats["mean"])])
    worksheet.append(["tiempo_derivado_por_correo_mediana_s", _seconds(per_email_stats["median"])])
    worksheet.append([])
    worksheet.append(["Detalle MCDM"])
    worksheet.append(
        [
            "mcdm_metric_id",
            "job_id",
            "batch_id",
            "operation",
            "status",
            "email_count",
            "started_at",
            "finished_at",
            "global_duration_s",
            "derived_duration_per_email_s",
            "derived_formula",
            "error_message",
        ]
    )
    for metric in metrics:
        worksheet.append(
            [
                str(metric.id),
                str(metric.job_id) if metric.job_id else None,
                str(metric.batch_id) if metric.batch_id else None,
                metric.operation,
                metric.status,
                metric.email_count,
                _utc_iso(metric.started_at),
                _utc_iso(metric.finished_at),
                _seconds(metric.duration_ms),
                _seconds(metric.duration_per_email_ms),
                "global_duration_s / email_count",
                metric.error_message,
            ]
        )
    worksheet.freeze_panes = "A16"
    _style_export_worksheet(worksheet)


def _weight_by_subcriterion_key() -> Dict[str, float]:
    return {
        str(item["key"]): float(item["weight"])
        for item in MCDM_SUBCRITERIA_WEIGHTS
    }


def _range_anchor_text(definition: Dict[str, Any]) -> Optional[str]:
    piecewise = definition.get("mcdm_piecewise")
    if not isinstance(piecewise, dict):
        return None
    anchors = piecewise.get("anchors")
    if not isinstance(anchors, list):
        return None
    parts: List[str] = []
    for anchor in anchors:
        if not isinstance(anchor, dict):
            continue
        parts.append(f'{anchor.get("value")} -> {anchor.get("score")}')
    return "; ".join(parts)


def _append_weights_and_ranges_export_sheet(
    workbook: Any,
    definitions: Sequence[Dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet(title="Pesos y rangos")
    raw_weights = _weight_by_subcriterion_key()
    raw_weight_total = sum(raw_weights.values())

    worksheet.append(["Configuracion enlaces", None])
    worksheet.append(["max_link_api_lookups_per_email", int(MAX_LINK_API_LOOKUPS_PER_EMAIL)])
    worksheet.append(["link_api_skip_threshold", int(LINK_API_SKIP_THRESHOLD)])
    worksheet.append(["link_api_skip_reason", LINK_API_SKIP_REASON])
    worksheet.append([])

    worksheet.append(["Pesos MCDM"])
    worksheet.append(
        [
            "subcriterion_key",
            "label",
            "family",
            "vector_field",
            "objective",
            "raw_weight",
            "raw_weight_total",
            "normalized_weight",
            "normalized_weight_percent",
            "value_type",
            "measurement",
        ]
    )
    for definition in definitions:
        key = str(definition["key"])
        normalized_weight = float(definition.get("mcdm_weight") or 0.0)
        worksheet.append(
            [
                key,
                definition.get("label"),
                definition.get("family"),
                definition.get("vector_field"),
                definition.get("mcdm_objective"),
                raw_weights.get(key, 0.0),
                raw_weight_total,
                normalized_weight,
                normalized_weight * 100.0,
                definition.get("value_type"),
                definition.get("measurement"),
            ]
        )

    worksheet.append([])
    worksheet.append(["Rangos MCDM"])
    worksheet.append(
        [
            "subcriterion_key",
            "label",
            "vector_field",
            "method",
            "anchor_index",
            "value",
            "score",
            "all_anchors",
        ]
    )
    for definition in definitions:
        key = str(definition["key"])
        piecewise = definition.get("mcdm_piecewise")
        anchor_text = _range_anchor_text(definition)
        if not isinstance(piecewise, dict):
            worksheet.append(
                [
                    key,
                    definition.get("label"),
                    definition.get("vector_field"),
                    None,
                    None,
                    None,
                    None,
                    None,
                ]
            )
            continue
        anchors = piecewise.get("anchors")
        if not isinstance(anchors, list):
            continue
        for index, anchor in enumerate(anchors, start=1):
            if not isinstance(anchor, dict):
                continue
            worksheet.append(
                [
                    key,
                    definition.get("label"),
                    definition.get("vector_field"),
                    piecewise.get("method"),
                    index,
                    anchor.get("value"),
                    anchor.get("score"),
                    anchor_text,
                ]
            )

    worksheet.append([])
    worksheet.append(["Mapeos SPF/DKIM/DMARC"])
    worksheet.append(["subcriterion", "source", "value", "score", "notes"])
    for value, score in sorted(SPF_SCORE_MAP.items()):
        worksheet.append(["spf", "SPF_SCORE_MAP", value, score, None])
    for value, score in sorted(DKIM_SCORE_MAP.items()):
        worksheet.append(["dkim", "DKIM_SCORE_MAP", value, score, None])
    for value, score in sorted(DMARC_STATUS_SCORE_MAP.items()):
        worksheet.append(["dmarc", "DMARC_STATUS_SCORE_MAP", value, score, None])
    for value, score in sorted(DMARC_FAIL_POLICY_SCORE_MAP.items()):
        worksheet.append(["dmarc", "DMARC_FAIL_POLICY_SCORE_MAP", value, score, "score usado cuando DMARC falla con esa politica"])
    worksheet.append(
        [
            "dmarc",
            "DMARC_AUTH_ALIGNMENT_RULE_VERSION",
            DMARC_AUTH_ALIGNMENT_RULE_VERSION,
            None,
            "version de la regla de ajuste por contexto SPF/DKIM",
        ]
    )
    worksheet.append(
        [
            "dmarc",
            "DMARC_PASS_WITH_AUTH_FAILURE_SCORE",
            "dmarc=pass with SPF or DKIM not pass",
            DMARC_PASS_WITH_AUTH_FAILURE_SCORE,
            "score minimo aplicado por la regla de contexto SPF/DKIM",
        ]
    )
    worksheet.freeze_panes = "A7"
    _style_export_worksheet(worksheet)


def _append_api_call_export_sheet(
    workbook: Any,
    batch: Batch,
    metrics: Sequence[EmailAnalysisMetric],
    api_calls: Sequence[ApiCallMetric],
) -> None:
    worksheet = workbook.create_sheet(title="Llamadas API")
    email_by_id = {str(email.id): email for email in _processable_batch_emails(batch)}
    metric_by_id = {str(metric.id): metric for metric in metrics}
    durations = [float(call.duration_ms) for call in api_calls]
    duration_stats = _stats(durations)

    worksheet.append(["Resumen", None])
    worksheet.append(["llamadas_total", len(api_calls)])
    worksheet.append(["duracion_total_s", _seconds(duration_stats["total"])])
    worksheet.append(["duracion_min_s", _seconds(duration_stats["min"])])
    worksheet.append(["duracion_max_s", _seconds(duration_stats["max"])])
    worksheet.append(["duracion_media_s", _seconds(duration_stats["mean"])])
    worksheet.append(["duracion_mediana_s", _seconds(duration_stats["median"])])
    worksheet.append([])
    worksheet.append(["Resumen por endpoint"])
    worksheet.append(["provider", "endpoint", "method", "llamadas", "errores", "duracion_total_s", "duracion_media_s", "duracion_mediana_s"])

    endpoint_groups: Dict[Tuple[str, str, str], List[ApiCallMetric]] = {}
    for call in api_calls:
        key = (call.provider, call.endpoint, call.method)
        endpoint_groups.setdefault(key, []).append(call)
    for (provider, endpoint, method), group in sorted(endpoint_groups.items()):
        group_durations = [float(call.duration_ms) for call in group]
        group_stats = _stats(group_durations)
        worksheet.append(
            [
                provider,
                endpoint,
                method,
                len(group),
                sum(1 for call in group if call.success is False or call.error),
                _seconds(group_stats["total"]),
                _seconds(group_stats["mean"]),
                _seconds(group_stats["median"]),
            ]
        )

    worksheet.append([])
    worksheet.append(["Detalle de llamadas"])
    worksheet.append(
        [
            "api_call_id",
            "analysis_id",
            "job_id",
            "job_type",
            "email_id",
            "email_name",
            "provider",
            "endpoint",
            "method",
            "target",
            "http_status",
            "success",
            "duration_s",
            "error",
            "created_at",
        ]
    )
    for call in api_calls:
        email = email_by_id.get(str(call.email_id))
        metric = metric_by_id.get(str(call.analysis_metric_id))
        worksheet.append(
            [
                str(call.id),
                str(call.analysis_metric_id),
                str(call.job_id) if call.job_id else None,
                metric.job_type if metric else None,
                str(call.email_id),
                email.name if email else None,
                call.provider,
                call.endpoint,
                call.method,
                call.target,
                call.http_status,
                call.success,
                _seconds(call.duration_ms),
                call.error,
                _utc_iso(call.created_at),
            ]
        )
    _style_export_worksheet(worksheet)


def serialize_email_summary(db: Session, email: Email) -> Dict[str, Any]:
    latest_job = _latest_job_for_email(db, email)
    return {
        "id": str(email.id),
        "name": email.name,
        "subject": email.subject,
        "created_at": _utc_iso(email.created_at),
        "batch": {
            "id": str(email.batch.id),
            "name": email.batch.name,
        }
        if email.batch
        else None,
        "selected_subcriteria": _effective_selected_subcriteria(email),
        "status": _email_status(db, email),
        "mcdm_score": email.enrichment.mcdm_score if email.enrichment else None,
        "mcdm_is_mock": email.enrichment.mcdm_is_mock if email.enrichment else True,
        "mcdm_method": email.enrichment.mcdm_method if email.enrichment else None,
        "processing_error": _email_processing_error_payload(email),
        "job": _serialize_job(latest_job),
    }


def serialize_email_detail(db: Session, email: Email) -> Dict[str, Any]:
    latest_job = _latest_job_for_email(db, email)
    headers = (email.raw_json or {}).get("headers") or {}
    metadata = (email.raw_json or {}).get("metadata") or {}
    missing_subcriteria = _missing_subcriteria_for_email(email)
    return {
        **serialize_email_summary(db, email),
        "metadata": metadata,
        "headers_summary": {
            "from": headers.get("from"),
            "to": headers.get("to"),
            "subject": headers.get("subject") or email.subject,
            "date": headers.get("date"),
            "message_id": headers.get("message_id"),
            "return_path": headers.get("return_path"),
        },
        "vector": {
            "version": email.enrichment.vector.vector_version if email.enrichment and email.enrichment.vector else None,
            "by_key": _model_to_vector_dict(email.enrichment.vector if email.enrichment else None),
        },
        "missing_subcriteria": missing_subcriteria,
        "subcriteria": _serialize_subcriteria_summary(email),
        "job": _serialize_job(latest_job),
    }


def serialize_batch_summary(db: Session, batch: Batch) -> Dict[str, Any]:
    latest_job = _latest_job_for_batch(db, batch.id)
    processable_emails = _processable_batch_emails(batch)
    active_email_jobs = [
        job
        for email in batch.emails
        for job in email.jobs
        if _job_is_active(job)
    ]
    discarded = sum(1 for email in batch.emails if _email_is_discarded(email))
    completed = sum(1 for email in processable_emails if _selected_complete(email)[0])
    mcdm_scores = [
        email.enrichment.mcdm_score
        for email in processable_emails
        if email.enrichment is not None and email.enrichment.mcdm_score is not None
    ]
    if _job_is_active(latest_job):
        status = latest_job.status
    elif active_email_jobs:
        status = max(
            active_email_jobs,
            key=lambda job: (
                job.created_at or datetime.min,
                str(job.id),
            ),
        ).status
    elif latest_job is not None and latest_job.status == "failed":
        status = "failed"
    elif discarded > 0:
        status = "failed"
    elif processable_emails and completed == len(processable_emails):
        status = "completed"
    else:
        status = "partial"
    return {
        "id": str(batch.id),
        "name": batch.name,
        "created_at": _utc_iso(batch.created_at),
        "total_emails": batch.total_emails,
        "processable_emails": len(processable_emails),
        "discarded_emails": discarded,
        "completed_emails": completed,
        "mcdm_score": max(mcdm_scores) if mcdm_scores else None,
        "status": status,
        "job": _serialize_job(latest_job),
    }


def _email_mcdm_desc_key(email: Email) -> Tuple[bool, float, datetime, str]:
    score = email.enrichment.mcdm_score if email.enrichment is not None else None
    return (
        score is None,
        -(score or 0.0),
        email.created_at or datetime.min,
        str(email.id),
    )


def serialize_batch_detail(
    db: Session,
    batch: Batch,
    *,
    email_limit: Optional[int] = None,
    email_offset: int = 0,
) -> Dict[str, Any]:
    selected = sorted({key for email in _processable_batch_emails(batch) for key in _effective_selected_subcriteria(email)})
    missing_subcriteria = _missing_subcriteria_for_batch(batch)
    emails = sorted(batch.emails, key=_email_mcdm_desc_key)
    if email_limit is not None:
        emails = emails[email_offset : email_offset + email_limit]
    return {
        **serialize_batch_summary(db, batch),
        "selected_subcriteria": selected,
        "missing_subcriteria": missing_subcriteria,
        "emails": [serialize_email_summary(db, email) for email in emails],
        "subcriteria": [
            {
                "key": definition["key"],
                "label": definition["label"],
                "family": definition["family"],
                "vector_field": definition["vector_field"],
                "value_type": definition["value_type"],
                "mcdm_objective": definition["mcdm_objective"],
                "mcdm_weight": definition["mcdm_weight"],
                "measurement": definition["measurement"],
            }
            for definition in available_subcriteria()
            if definition["key"] in selected
        ],
    }


def _round_value(value: float) -> float:
    return round(float(value), 6)


def _percentile(sorted_values: Sequence[float], ratio: float) -> float:
    if not sorted_values:
        return 0.0
    if len(sorted_values) == 1:
        return sorted_values[0]

    position = (len(sorted_values) - 1) * ratio
    lower_index = int(position)
    upper_index = min(lower_index + 1, len(sorted_values) - 1)
    lower_value = sorted_values[lower_index]
    upper_value = sorted_values[upper_index]
    interpolation = position - lower_index
    return lower_value + ((upper_value - lower_value) * interpolation)


def _numeric_stats(values: Sequence[float]) -> Optional[Dict[str, float]]:
    if not values:
        return None

    sorted_values = sorted(values)
    mean = sum(sorted_values) / len(sorted_values)
    return {
        "min": _round_value(sorted_values[0]),
        "q1": _round_value(_percentile(sorted_values, 0.25)),
        "median": _round_value(_percentile(sorted_values, 0.5)),
        "mean": _round_value(mean),
        "q3": _round_value(_percentile(sorted_values, 0.75)),
        "max": _round_value(sorted_values[-1]),
    }


def _build_numeric_bins(values: Sequence[float]) -> List[Dict[str, Any]]:
    if not values:
        return []

    total = len(values)
    unique_values = sorted(set(values))
    if len(unique_values) <= 6:
        return [
            {
                "label": str(_round_value(value)),
                "start": value,
                "end": value,
                "center": _round_value(value),
                "count": values.count(value),
                "ratio": _round_value(values.count(value) / total),
            }
            for value in unique_values
        ]

    min_value = min(values)
    max_value = max(values)
    if min_value == max_value:
        return [
            {
                "label": str(_round_value(min_value)),
                "start": min_value,
                "end": max_value,
                "center": _round_value(min_value),
                "count": len(values),
                "ratio": 1.0,
            }
        ]

    bucket_count = 5
    step = (max_value - min_value) / bucket_count
    bins: List[Dict[str, Any]] = []
    for index in range(bucket_count):
        start = min_value + index * step
        end = max_value if index == bucket_count - 1 else min_value + (index + 1) * step
        count = 0
        for value in values:
            if index == bucket_count - 1:
                if start <= value <= end:
                    count += 1
            else:
                if start <= value < end:
                    count += 1
        bins.append(
            {
                "label": f"{_round_value(start)}-{_round_value(end)}",
                "start": _round_value(start),
                "end": _round_value(end),
                "center": _round_value((start + end) / 2),
                "count": count,
                "ratio": _round_value(count / total),
            }
        )
    return bins


def _segment_label(subcriterion_key: str, raw_value: float) -> str:
    semantic_labels = SEMANTIC_VALUE_LABELS_BY_SUBCRITERION_KEY.get(subcriterion_key, {})
    rounded_value = _round_value(raw_value)
    if rounded_value in semantic_labels:
        return semantic_labels[rounded_value]
    return str(rounded_value)


def _numeric_segments(subcriterion_key: str, values: Sequence[float]) -> List[Dict[str, Any]]:
    if not values:
        return []

    counts: Dict[float, int] = {}
    for value in values:
        rounded = _round_value(value)
        counts[rounded] = counts.get(rounded, 0) + 1

    total = len(values)
    segments: List[Dict[str, Any]] = []
    for raw_value in sorted(counts):
        count = counts[raw_value]
        segments.append(
            {
                "label": _segment_label(subcriterion_key, raw_value),
                "value": raw_value,
                "count": count,
                "ratio": _round_value(count / total),
            }
        )
    return segments


def _numeric_points(subcriterion_key: str, values: Sequence[float]) -> List[Dict[str, Any]]:
    return _numeric_segments(subcriterion_key, values)


def _bucket_values(values: List[Any]) -> List[Dict[str, Any]]:
    if not values:
        return []

    unique = sorted(set(values))
    if all(isinstance(value, int) and value in (0, 1) for value in unique):
        return [{"label": str(value), "count": values.count(value)} for value in unique]

    if len(unique) <= 8:
        return [{"label": str(value), "count": values.count(value)} for value in unique]

    numeric = [float(value) for value in values if isinstance(value, (int, float))]
    if not numeric:
        return [{"label": str(value), "count": values.count(value)} for value in unique]

    min_value = min(numeric)
    max_value = max(numeric)
    if min_value == max_value:
        return [{"label": str(round(min_value, 3)), "count": len(numeric)}]

    bucket_count = 5
    step = (max_value - min_value) / bucket_count
    buckets: List[Dict[str, Any]] = []
    for index in range(bucket_count):
        start = min_value + index * step
        end = max_value if index == bucket_count - 1 else min_value + (index + 1) * step
        count = 0
        for value in numeric:
            if index == bucket_count - 1:
                if start <= value <= end:
                    count += 1
            else:
                if start <= value < end:
                    count += 1
        label = f"{round(start, 3)} - {round(end, 3)}"
        buckets.append({"label": label, "count": count})
    return buckets


def batch_score_analytics_overview(batch: Batch) -> Dict[str, Any]:
    processable_emails = _processable_batch_emails(batch)
    selected = sorted({key for email in processable_emails for key in _effective_selected_subcriteria(email)})
    definitions = [
        definition
        for definition in available_subcriteria()
        if definition["key"] in selected
    ]

    items: List[Dict[str, Any]] = []
    emails_total = len(processable_emails)
    for definition in definitions:
        chart_config = chart_config_for_subcriterion(definition["key"])
        chart_type = chart_config["chart_type"]
        display_mode = chart_config["display_mode"]
        records: List[Dict[str, Any]] = []
        numeric_values: List[float] = []
        labels: List[str] = []

        for email in processable_emails:
            enrichment = email.enrichment
            raw_result = getattr(enrichment, definition["enrichment_column"]) if enrichment is not None else None
            normalized_result = _normalize_stored_subcriterion_result(definition["key"], raw_result)
            display_value = _display_subcriterion_value(definition["key"], normalized_result)
            record = extract_chart_display_record(definition["key"], normalized_result, display_value)
            if record is None:
                continue
            records.append(record)
            label = record.get("label")
            if isinstance(label, str) and label:
                labels.append(label)
            numeric_value = record.get("numeric_value")
            if isinstance(numeric_value, (int, float)):
                numeric_values.append(_round_value(float(numeric_value)))

        emails_with_value = len(records)
        bins: List[Dict[str, Any]] = []
        segments: List[Dict[str, Any]] = []
        points: List[Dict[str, Any]] = []

        if chart_type == "band_bars":
            if display_mode == "native_score_band":
                bins = build_score_band_bins(numeric_values)
            else:
                bins = build_label_bins(definition["key"], labels)
        elif chart_type == "histogram":
            bins = build_semantic_histogram_bins(definition["key"], numeric_values)
        elif chart_type == "bubble_lane":
            points = build_numeric_points(numeric_values)
        elif chart_type in {"stacked_bar", "waffle", "donut", "pie"}:
            segments = build_label_segments(definition["key"], labels)

        items.append(
            {
                "subcriterion": {
                    "key": definition["key"],
                    "label": definition["label"],
                    "family": definition["family"],
                    "vector_field": definition["vector_field"],
                    "value_type": definition["value_type"],
                    "mcdm_objective": definition["mcdm_objective"],
                    "mcdm_weight": definition["mcdm_weight"],
                    "measurement": definition["measurement"],
                },
                "chart_type": chart_type,
                "value_source": display_mode,
                "emails_total": emails_total,
                "emails_with_value": emails_with_value,
                "emails_without_value": max(emails_total - emails_with_value, 0),
                "coverage_ratio": _round_value(emails_with_value / emails_total) if emails_total else 0.0,
                "value_stats": _numeric_stats(numeric_values) if numeric_values else None,
                "bins": bins,
                "segments": segments,
                "points": points,
            }
        )

    return {
        "batch_id": str(batch.id),
        "batch_name": batch.name,
        "items": items,
    }


def batch_subcriterion_analytics(batch: Batch, subcriterion_key: str) -> Dict[str, Any]:
    definition = get_subcriterion_definition(subcriterion_key)
    processable_emails = _processable_batch_emails(batch)
    emails_total = len(processable_emails)
    series: List[Dict[str, Any]] = []

    for vector_field in definition["vector_fields"]:
        values: List[Any] = []
        missing = 0
        for email in processable_emails:
            vector = email.enrichment.vector if email.enrichment and email.enrichment.vector else None
            if vector is None:
                missing += 1
                continue
            value = _normalize_vector_field_value(vector_field, getattr(vector, vector_field))
            if value is None:
                missing += 1
                continue
            values.append(value)

        series.append(
            {
                "metric_key": vector_field,
                "metric_label": VECTOR_FIELD_LABELS.get(vector_field, vector_field),
                "emails_total": emails_total,
                "emails_with_value": len(values),
                "emails_without_value": missing,
                "distribution": _bucket_values(values),
            }
        )

    return {
        "batch_id": str(batch.id),
        "batch_name": batch.name,
        "subcriterion": {
            "key": definition["key"],
            "label": definition["label"],
            "family": definition["family"],
            "vector_field": definition["vector_field"],
            "value_type": definition["value_type"],
            "mcdm_objective": definition["mcdm_objective"],
            "mcdm_weight": definition["mcdm_weight"],
            "measurement": definition["measurement"],
        },
        "series": series,
    }


def _create_empty_related(email: Email) -> None:
    email.enrichment = EmailEnrichment()
    email.enrichment.vector = EmailMcdmVector()


def _sanitize_json_for_postgres(value: Any) -> Any:
    if isinstance(value, str):
        return value.replace("\x00", "").encode("utf-8", errors="ignore").decode("utf-8")
    if isinstance(value, list):
        return [_sanitize_json_for_postgres(item) for item in value]
    if isinstance(value, dict):
        return {
            _sanitize_json_for_postgres(key) if isinstance(key, str) else key: _sanitize_json_for_postgres(item)
            for key, item in value.items()
        }
    return value


def _raw_upload_error_payload(
    *,
    file_name: str,
    index: Optional[int],
    raw_sha: str,
    raw_size: int,
    error_code: str,
    error_message: str,
) -> Dict[str, Any]:
    timestamp = datetime.utcnow()
    return {
        "metadata": {
            "file_name": file_name,
            "file_stem": Path(file_name).stem,
            "file_ext": Path(file_name).suffix,
            "file_path": None,
            "raw_size": raw_size,
            "raw_sha256": raw_sha,
            "parse_timestamp": _utc_iso(timestamp),
            "upload_index": index,
        },
        "headers": {},
        "parts": [],
        "body": {"text": None, "html": None},
        "urls": [],
        "attachments": [],
        "encodings": {
            "charsets_detected": [],
            "transfer_encodings": [],
        },
        "enrichment": {},
        "raw": {"headers": "", "body": ""},
        "processing_error": {
            "state": EMAIL_PROCESSING_STATE_ERROR,
            "code": error_code[:128],
            "message": error_message,
            "at": _utc_iso(timestamp),
        },
    }


def _create_upload_error_email(
    *,
    file_name: str,
    content: bytes,
    batch_id: UUID,
    index: Optional[int],
    error_code: str,
    error_message: str,
) -> Email:
    raw_sha = sha256_bytes(content)
    payload = _sanitize_json_for_postgres(
        _raw_upload_error_payload(
            file_name=file_name,
            index=index,
            raw_sha=raw_sha,
            raw_size=len(content),
            error_code=error_code,
            error_message=error_message,
        )
    )
    email = Email(
        name=Path(file_name).stem or file_name,
        subject=None,
        raw_json=payload,
        batch_id=batch_id,
        selected_subcriteria=[],
        processing_state=EMAIL_PROCESSING_STATE_ERROR,
        processing_error_code=error_code[:128],
        processing_error_message=error_message,
        processing_error_at=datetime.utcnow(),
    )
    _create_empty_related(email)
    return email


def _batch_has_any_active_jobs(db: Session, batch_id: UUID) -> bool:
    stmt = (
        select(Job.id)
        .outerjoin(Email, Job.email_id == Email.id)
        .where(
            Job.status.in_(("queued", "running")),
            or_(Job.batch_id == batch_id, Email.batch_id == batch_id),
        )
        .limit(1)
    )
    return db.scalar(stmt) is not None


def _load_batch_emails_for_mcdm(db: Session, batch_id: UUID) -> List[Email]:
    stmt = (
        select(Email)
        .where(
            Email.batch_id == batch_id,
            Email.processing_state.not_in(tuple(EMAIL_DISCARDED_STATES)),
        )
        .options(
            load_only(Email.id, Email.batch_id, Email.created_at, Email.raw_json, Email.processing_state),
            selectinload(Email.enrichment).selectinload(EmailEnrichment.vector),
        )
        .order_by(Email.created_at.asc(), Email.id.asc())
    )
    return list(db.scalars(stmt).unique())


def _recompute_mcdm_for_emails(
    emails: Sequence[Email],
    *,
    db: Optional[Session] = None,
    job: Optional[Job] = None,
    operation: str = "recompute_mcdm",
) -> None:
    if not emails:
        return

    started_at = datetime.utcnow()
    started_perf = time.perf_counter()
    batch_id = next((email.batch_id for email in emails if email.batch_id is not None), None)
    processed_payloads: List[Dict[str, Any]] = []
    try:
        for email in emails:
            raw_json = email.raw_json if isinstance(email.raw_json, dict) else {}
            numeric_values = raw_json.get("numeric_values")
            payload = dict(raw_json)
            if isinstance(numeric_values, dict):
                payload["numeric_values"] = dict(numeric_values)
            processed_payloads.append(payload)
        if len(processed_payloads) == 1 and emails[0].batch_id is None:
            mcdm_blocks = [compute_mcdm_block_for_email_with_hardcoded_references(processed_payloads[0])]
        else:
            mcdm_blocks = compute_mcdm_blocks_for_emails(processed_payloads)
        for email, processed_email, mcdm_block in zip(emails, processed_payloads, mcdm_blocks):
            _apply_mcdm_block(email, processed_email, mcdm_block)
        if db is not None:
            _persist_mcdm_runtime_metric(
                db,
                job=job,
                batch_id=batch_id,
                operation=operation,
                started_at=started_at,
                duration_ms=(time.perf_counter() - started_perf) * 1000,
                email_count=len(emails),
            )
    except Exception as exc:
        if db is not None:
            _persist_mcdm_runtime_metric(
                db,
                job=job,
                batch_id=batch_id,
                operation=operation,
                started_at=started_at,
                duration_ms=(time.perf_counter() - started_perf) * 1000,
                email_count=len(emails),
                status="failed",
                error_message=str(exc),
            )
        raise


def _apply_mcdm_block(
    email: Email,
    processed_email: Dict[str, Any],
    mcdm_block: Dict[str, Any],
) -> None:
    raw_json = dict(email.raw_json or {})
    raw_json["mcdm"] = dict(mcdm_block)
    email.raw_json = raw_json

    payload = dict(processed_email)
    payload["mcdm"] = dict(mcdm_block)
    _apply_processed_email(email, payload)


def _apply_processed_email(email: Email, processed_email: Dict[str, Any]) -> None:
    enrichment_payload = processed_email.get("enrichment") or {}
    numeric_values = ((processed_email.get("numeric_values") or {}).get("by_key")) or {}
    mcdm = processed_email.get("mcdm") or {}

    email.processing_state = EMAIL_PROCESSING_STATE_READY
    email.processing_error_code = None
    email.processing_error_message = None
    email.processing_error_at = None

    if email.enrichment is None:
        email.enrichment = EmailEnrichment()
    if email.enrichment.vector is None:
        email.enrichment.vector = EmailMcdmVector()

    for key, column in ENRICHMENT_JSON_COLUMNS.items():
        if key in enrichment_payload:
            normalized_result = _normalize_stored_subcriterion_result(
                key,
                enrichment_payload.get(key),
            )
            setattr(email.enrichment, column, normalized_result)

    email.enrichment.mcdm_score = mcdm.get("score")
    email.enrichment.mcdm_is_mock = bool(mcdm.get("is_mock", True))
    email.enrichment.mcdm_method = mcdm.get("method")
    email.enrichment.updated_at = datetime.utcnow()

    for field in ALL_VECTOR_FIELDS:
        if field in numeric_values:
            setattr(
                email.enrichment.vector,
                field,
                _normalize_vector_field_value(field, numeric_values.get(field)),
            )
    version = (processed_email.get("numeric_values") or {}).get("version")
    if isinstance(version, int):
        email.enrichment.vector.vector_version = version
    email.enrichment.vector.updated_at = datetime.utcnow()


async def _read_upload_file(upload: UploadFile, *, max_file_bytes: int) -> Tuple[str, bytes]:
    chunks: List[bytes] = []
    total_bytes = 0

    while True:
        chunk = await upload.read(UPLOAD_READ_CHUNK_BYTES)
        if not chunk:
            break
        total_bytes += len(chunk)
        if total_bytes > max_file_bytes:
            raise ValueError(
                f'El archivo "{upload.filename or "email.eml"}" supera el limite de {max_file_bytes} bytes.'
            )
        chunks.append(chunk)

    return upload.filename or "email.eml", b"".join(chunks)


async def read_uploads(
    files: Sequence[UploadFile],
    *,
    max_file_bytes: int,
    max_total_bytes: int,
) -> List[Tuple[str, bytes]]:
    items: List[Tuple[str, bytes]] = []
    total_bytes = 0
    for upload in files:
        item = await _read_upload_file(upload, max_file_bytes=max_file_bytes)
        total_bytes += len(item[1])
        if total_bytes > max_total_bytes:
            raise ValueError(f"El total subido supera el limite de {max_total_bytes} bytes.")
        items.append(item)
    return items


def _email_raw_size_bytes(raw_json: Any) -> int:
    if not isinstance(raw_json, dict):
        return 0
    metadata = raw_json.get("metadata")
    if not isinstance(metadata, dict):
        return 0
    raw_size = metadata.get("raw_size")
    if isinstance(raw_size, bool):
        return 0
    if isinstance(raw_size, (int, float)):
        return max(int(raw_size), 0)
    try:
        return max(int(str(raw_size)), 0)
    except (TypeError, ValueError):
        return 0


def _batch_raw_size_bytes(db: Session, batch_id: UUID) -> int:
    rows = db.scalars(select(Email.raw_json).where(Email.batch_id == batch_id)).all()
    return sum(_email_raw_size_bytes(raw_json) for raw_json in rows)


async def create_batch_upload_from_files(
    db: Session,
    background_tasks: BackgroundTasks,
    batch_name: str,
    files: Sequence[UploadFile],
    selected_subcriteria: Sequence[str],
    *,
    max_file_bytes: int,
    max_total_bytes: int,
) -> Dict[str, Any]:
    normalized_selected = list(normalize_selected_subcriteria(list(selected_subcriteria)))
    batch = Batch(name=batch_name, total_emails=len(files))
    db.add(batch)
    db.flush()

    email_ids: List[str] = []
    total_bytes = 0
    for index, upload in enumerate(files, start=1):
        file_name, content = await _read_upload_file(upload, max_file_bytes=max_file_bytes)
        total_bytes += len(content)
        if total_bytes > max_total_bytes:
            raise ValueError(f"El total subido supera el limite de {max_total_bytes} bytes.")

        try:
            parsed_email = _sanitize_json_for_postgres(
                parse_uploaded_email(file_name=file_name, content=content, index=index)
            )
            email = Email(
                name=((parsed_email.get("metadata") or {}).get("file_stem") or file_name),
                subject=((parsed_email.get("headers") or {}).get("subject")),
                raw_json=parsed_email,
                batch_id=batch.id,
                selected_subcriteria=normalized_selected,
            )
        except Exception as exc:
            email = _create_upload_error_email(
                file_name=file_name,
                content=content,
                batch_id=batch.id,
                index=index,
                error_code="upload_parse_error",
                error_message=f"No se pudo parsear o preparar el correo: {exc}",
            )
        _create_empty_related(email)
        db.add(email)
        db.flush()
        email_ids.append(str(email.id))

    job = Job(
        job_type="enrich_batch",
        target_type="batch",
        target_id=str(batch.id),
        batch_id=batch.id,
        selected_subcriteria=normalized_selected,
        status="queued",
        progress_current=0,
        progress_total=len(email_ids),
    )
    db.add(job)
    db.commit()
    db.refresh(batch)
    db.refresh(job)

    background_tasks.add_task(run_batch_job, str(job.id), email_ids, normalized_selected, False)

    return {
        "batch_id": str(batch.id),
        "job_id": str(job.id),
        "email_ids": email_ids,
    }


def create_pending_batch_upload(db: Session, batch_name: str) -> Dict[str, Any]:
    batch = Batch(name=batch_name, total_emails=0)
    db.add(batch)
    db.commit()
    db.refresh(batch)
    return {
        "batch_id": str(batch.id),
        "uploaded_emails": 0,
    }


async def append_batch_upload_files(
    db: Session,
    batch: Batch,
    files: Sequence[UploadFile],
    *,
    expected_start_index: int,
    max_file_bytes: int,
    max_batch_files: int,
    max_total_bytes: int,
) -> Dict[str, Any]:
    if _batch_has_any_active_jobs(db, batch.id):
        raise ValueError("No se pueden anadir correos a un lote con jobs activos.")
    if batch.total_emails + len(files) > max_batch_files:
        raise ValueError(f"El lote supera el maximo permitido de {max_batch_files} archivos.")
    if expected_start_index < 1:
        raise ValueError("El indice esperado de subida debe ser >= 1.")

    expected_previous = expected_start_index - 1
    expected_last = expected_previous + len(files)
    if batch.total_emails >= expected_last:
        return {
            "batch_id": str(batch.id),
            "uploaded_emails": batch.total_emails,
            "email_ids": [],
        }
    if batch.total_emails != expected_previous:
        raise ValueError(
            f"Estado de subida incoherente: el lote tiene {batch.total_emails} correos "
            f"y se esperaba continuar en {expected_start_index}."
        )

    existing_total_bytes = _batch_raw_size_bytes(db, batch.id)
    chunk_bytes = 0
    email_ids: List[str] = []
    try:
        for offset, upload in enumerate(files):
            file_name, content = await _read_upload_file(upload, max_file_bytes=max_file_bytes)
            chunk_bytes += len(content)
            if existing_total_bytes + chunk_bytes > max_total_bytes:
                raise ValueError(f"El lote supera el limite total de {max_total_bytes} bytes.")

            current_index = expected_start_index + offset
            try:
                parsed_email = _sanitize_json_for_postgres(
                    parse_uploaded_email(
                        file_name=file_name,
                        content=content,
                        index=current_index,
                    )
                )
                email = Email(
                    name=((parsed_email.get("metadata") or {}).get("file_stem") or file_name),
                    subject=((parsed_email.get("headers") or {}).get("subject")),
                    raw_json=parsed_email,
                    batch_id=batch.id,
                    selected_subcriteria=[],
                )
            except Exception as exc:
                email = _create_upload_error_email(
                    file_name=file_name,
                    content=content,
                    batch_id=batch.id,
                    index=current_index,
                    error_code="upload_parse_error",
                    error_message=f"No se pudo parsear o preparar el correo: {exc}",
                )
            _create_empty_related(email)
            db.add(email)
            db.flush()
            email_ids.append(str(email.id))

        batch.total_emails += len(email_ids)
        db.commit()
        db.refresh(batch)
    except Exception:
        db.rollback()
        raise

    return {
        "batch_id": str(batch.id),
        "uploaded_emails": batch.total_emails,
        "email_ids": email_ids,
    }


def finalize_batch_upload(
    db: Session,
    background_tasks: BackgroundTasks,
    batch: Batch,
    selected_subcriteria: Sequence[str],
) -> Dict[str, Any]:
    if _batch_has_any_active_jobs(db, batch.id):
        raise ValueError("No se puede finalizar un lote con jobs activos.")

    email_ids = [str(email.id) for email in sorted(batch.emails, key=lambda item: (item.created_at, str(item.id)))]
    if not email_ids:
        raise ValueError("No hay correos subidos en este lote.")

    normalized_selected = list(normalize_selected_subcriteria(list(selected_subcriteria)))
    for email in batch.emails:
        email.selected_subcriteria = list(normalized_selected)

    job = Job(
        job_type="enrich_batch",
        target_type="batch",
        target_id=str(batch.id),
        batch_id=batch.id,
        selected_subcriteria=normalized_selected,
        status="queued",
        progress_current=0,
        progress_total=len(email_ids),
    )
    db.add(job)
    db.commit()
    db.refresh(job)

    background_tasks.add_task(run_batch_job, str(job.id), email_ids, normalized_selected, False)

    return {
        "batch_id": str(batch.id),
        "job_id": str(job.id),
        "email_ids": email_ids,
    }


def create_single_upload(
    db: Session,
    background_tasks: BackgroundTasks,
    email_name: str,
    file_name: str,
    content: bytes,
    selected_subcriteria: Sequence[str],
) -> Dict[str, Any]:
    parsed_email = _sanitize_json_for_postgres(parse_uploaded_email(file_name=file_name, content=content))
    email = Email(
        name=email_name,
        subject=((parsed_email.get("headers") or {}).get("subject")),
        raw_json=parsed_email,
        selected_subcriteria=list(normalize_selected_subcriteria(list(selected_subcriteria))),
    )
    _create_empty_related(email)
    db.add(email)
    db.flush()

    job = Job(
        job_type="enrich_email",
        target_type="email",
        target_id=str(email.id),
        email_id=email.id,
        selected_subcriteria=list(email.selected_subcriteria),
        status="queued",
        progress_current=0,
        progress_total=len(email.selected_subcriteria),
    )
    db.add(job)
    db.commit()
    db.refresh(email)
    db.refresh(job)

    background_tasks.add_task(run_email_job, str(job.id), str(email.id), list(email.selected_subcriteria), False)

    return {
        "email_id": str(email.id),
        "job_id": str(job.id),
    }


def create_batch_upload(
    db: Session,
    background_tasks: BackgroundTasks,
    batch_name: str,
    files: Sequence[Tuple[str, bytes]],
    selected_subcriteria: Sequence[str],
) -> Dict[str, Any]:
    normalized_selected = list(normalize_selected_subcriteria(list(selected_subcriteria)))
    batch = Batch(name=batch_name, total_emails=len(files))
    db.add(batch)
    db.flush()

    email_ids: List[str] = []
    for file_name, content in files:
        parsed_email = _sanitize_json_for_postgres(parse_uploaded_email(file_name=file_name, content=content))
        email = Email(
            name=((parsed_email.get("metadata") or {}).get("file_stem") or file_name),
            subject=((parsed_email.get("headers") or {}).get("subject")),
            raw_json=parsed_email,
            batch_id=batch.id,
            selected_subcriteria=normalized_selected,
        )
        _create_empty_related(email)
        db.add(email)
        db.flush()
        email_ids.append(str(email.id))

    job = Job(
        job_type="enrich_batch",
        target_type="batch",
        target_id=str(batch.id),
        batch_id=batch.id,
        selected_subcriteria=normalized_selected,
        status="queued",
        progress_current=0,
        progress_total=len(email_ids),
    )
    db.add(job)
    db.commit()
    db.refresh(batch)
    db.refresh(job)

    background_tasks.add_task(run_batch_job, str(job.id), email_ids, normalized_selected, False)

    return {
        "batch_id": str(batch.id),
        "job_id": str(job.id),
        "email_ids": email_ids,
    }


def run_email_job(job_id: str, email_id: str, selected_subcriteria: Sequence[str], force: bool) -> None:
    session_factory = get_session_local()
    db = session_factory()
    analysis_started_at: Optional[datetime] = None
    analysis_started_perf: Optional[float] = None
    analysis_duration_ms: Optional[float] = None
    api_calls: List[Dict[str, Any]] = []
    try:
        job = db.get(Job, UUID(job_id))
        email = db.get(Email, UUID(email_id))
        if job is None or email is None:
            return
        job.status = "running"
        job.started_at = datetime.utcnow()
        db.commit()

        analysis_started_at = datetime.utcnow()
        analysis_started_perf = time.perf_counter()
        processed_email = dict(email.raw_json or {})

        def persist_subcriterion_progress(
            current_data: Dict[str, Any],
            _criterion_key: str,
            completed_count: int,
            total_count: int,
        ) -> None:
            email.raw_json = dict(current_data)
            _apply_processed_email(email, current_data)
            job.progress_total = total_count
            job.progress_current = completed_count
            db.commit()

        with collect_api_calls() as collected_api_calls:
            api_calls = collected_api_calls
            result = enrich_email_in_data(
                processed_email,
                selected_subcriteria=list(selected_subcriteria),
                force=force,
                include_numeric_values=True,
                on_subcriterion_completed=persist_subcriterion_progress,
            )
        api_calls = list(collected_api_calls)
        analysis_duration_ms = (time.perf_counter() - analysis_started_perf) * 1000
        email.raw_json = dict(processed_email)
        if email.batch_id:
            batch_emails = _load_batch_emails_for_mcdm(db, email.batch_id)
            _recompute_mcdm_for_emails(
                batch_emails,
                db=db,
                job=job,
                operation="analysis_batch_mcdm_refresh",
            )
        else:
            _recompute_mcdm_for_emails(
                [email],
                db=db,
                job=job,
                operation="analysis_email_mcdm_refresh",
            )
        job.progress_current = job.progress_total
        job.status = "completed"
        job.finished_at = datetime.utcnow()
        job.error_message = None
        if result["selected_subcriteria"]:
            email.selected_subcriteria = _merge_selected_subcriteria(
                email.selected_subcriteria,
                result["selected_subcriteria"],
            )
        _persist_email_analysis_metric(
            db,
            job=job,
            email=email,
            started_at=analysis_started_at,
            duration_ms=analysis_duration_ms,
            status="completed",
            selected_subcriteria=list(selected_subcriteria),
            api_calls=api_calls,
        )
        db.commit()
    except FatalExternalApiError as exc:
        if analysis_started_perf is not None:
            analysis_duration_ms = (time.perf_counter() - analysis_started_perf) * 1000
        db.rollback()
        job = db.get(Job, UUID(job_id))
        email = db.get(Email, UUID(email_id))
        if job is not None and email is not None:
            _mark_email_as_processing_cancelled(
                email,
                error_code=exc.error_code,
                error_message=f"Analisis cancelado por API externa en {exc.criterion_key}: {exc.error_message}",
            )
            job.status = "failed"
            job.error_message = email.processing_error_message
            job.finished_at = datetime.utcnow()
            if analysis_started_at is not None:
                _persist_email_analysis_metric(
                    db,
                    job=job,
                    email=email,
                    started_at=analysis_started_at,
                    duration_ms=analysis_duration_ms or 0.0,
                    status="failed",
                    selected_subcriteria=list(selected_subcriteria),
                    api_calls=api_calls,
                    error_message=email.processing_error_message,
                )
            db.commit()
    except Exception as exc:
        if analysis_started_perf is not None:
            analysis_duration_ms = (time.perf_counter() - analysis_started_perf) * 1000
        db.rollback()
        job = db.get(Job, UUID(job_id))
        email = db.get(Email, UUID(email_id))
        if job is not None and email is not None:
            _mark_email_as_processing_cancelled(
                email,
                error_code="processing_exception",
                error_message=f"Analisis cancelado por error interno: {exc}",
            )
            job.status = "failed"
            job.error_message = email.processing_error_message
            job.finished_at = datetime.utcnow()
            if analysis_started_at is not None:
                _persist_email_analysis_metric(
                    db,
                    job=job,
                    email=email,
                    started_at=analysis_started_at,
                    duration_ms=analysis_duration_ms or 0.0,
                    status="failed",
                    selected_subcriteria=list(selected_subcriteria),
                    api_calls=api_calls,
                    error_message=email.processing_error_message,
                )
            db.commit()
        elif job is not None:
            job.status = "failed"
            job.error_message = str(exc)
            job.finished_at = datetime.utcnow()
            db.commit()
    finally:
        db.close()


def _run_batch_email_job(
    *,
    job_id: str,
    email_id: str,
    selected_subcriteria: Sequence[str],
    force: bool,
) -> Dict[str, Any]:
    session_factory = get_session_local()
    analysis_started_at = datetime.utcnow()
    analysis_started_perf = time.perf_counter()
    api_calls: List[Dict[str, Any]] = []

    db = session_factory()
    try:
        job = db.get(Job, UUID(job_id))
        email = db.get(Email, UUID(email_id))
        if job is None:
            return {"email_id": email_id, "status": "job_missing", "error_message": "Job no encontrado"}
        if email is None:
            return {"email_id": email_id, "status": "email_missing", "error_message": "Correo no encontrado"}
        processed_email = dict(email.raw_json or {})
    finally:
        db.close()

    try:
        with collect_api_calls() as collected_api_calls:
            api_calls = collected_api_calls
            result = enrich_email_in_data(
                processed_email,
                selected_subcriteria=list(selected_subcriteria),
                force=force,
                include_numeric_values=True,
            )
        analysis_duration_ms = (time.perf_counter() - analysis_started_perf) * 1000

        db = session_factory()
        try:
            job = db.get(Job, UUID(job_id))
            email = db.get(Email, UUID(email_id))
            if job is None:
                return {"email_id": email_id, "status": "job_missing", "error_message": "Job no encontrado"}
            if email is None:
                return {"email_id": email_id, "status": "email_missing", "error_message": "Correo no encontrado"}
            email.raw_json = dict(processed_email)
            _apply_processed_email(email, processed_email)
            if result["selected_subcriteria"]:
                email.selected_subcriteria = _merge_selected_subcriteria(
                    email.selected_subcriteria,
                    result["selected_subcriteria"],
                )
            _persist_email_analysis_metric(
                db,
                job=job,
                email=email,
                started_at=analysis_started_at,
                duration_ms=analysis_duration_ms,
                status="completed",
                selected_subcriteria=list(selected_subcriteria),
                api_calls=list(api_calls),
            )
            db.commit()
            return {"email_id": email_id, "status": "completed", "error_message": None}
        finally:
            db.close()
    except FatalExternalApiError as exc:
        analysis_duration_ms = (time.perf_counter() - analysis_started_perf) * 1000
        db = session_factory()
        try:
            job = db.get(Job, UUID(job_id))
            email = db.get(Email, UUID(email_id))
            if job is None:
                return {"email_id": email_id, "status": "job_missing", "error_message": "Job no encontrado"}
            error_message = f"Analisis cancelado por API externa en {exc.criterion_key}: {exc.error_message}"
            if email is not None:
                _mark_email_as_processing_cancelled(
                    email,
                    error_code=exc.error_code,
                    error_message=error_message,
                )
                _persist_email_analysis_metric(
                    db,
                    job=job,
                    email=email,
                    started_at=analysis_started_at,
                    duration_ms=analysis_duration_ms,
                    status="failed",
                    selected_subcriteria=list(selected_subcriteria),
                    api_calls=list(api_calls),
                    error_message=email.processing_error_message,
                )
            db.commit()
            return {"email_id": email_id, "status": "failed_external_api", "error_message": error_message}
        finally:
            db.close()
    except Exception as exc:
        analysis_duration_ms = (time.perf_counter() - analysis_started_perf) * 1000
        db = session_factory()
        try:
            job = db.get(Job, UUID(job_id))
            email = db.get(Email, UUID(email_id))
            error_message = f"Analisis cancelado por error interno: {exc}"
            if job is None:
                return {"email_id": email_id, "status": "job_missing", "error_message": "Job no encontrado"}
            if email is not None:
                _mark_email_as_processing_cancelled(
                    email,
                    error_code="processing_exception",
                    error_message=error_message,
                )
                _persist_email_analysis_metric(
                    db,
                    job=job,
                    email=email,
                    started_at=analysis_started_at,
                    duration_ms=analysis_duration_ms,
                    status="failed",
                    selected_subcriteria=list(selected_subcriteria),
                    api_calls=list(api_calls),
                    error_message=error_message,
                )
            db.commit()
            return {"email_id": email_id, "status": "failed_internal", "error_message": error_message}
        finally:
            db.close()


def run_batch_job(job_id: str, email_ids: Sequence[str], selected_subcriteria: Sequence[str], force: bool) -> None:
    session_factory = get_session_local()
    db = session_factory()
    try:
        job = db.get(Job, UUID(job_id))
        if job is None:
            return
        job.status = "running"
        job.started_at = datetime.utcnow()
        db.commit()

        completed_count = 0
        failed_internal_messages: List[str] = []
        max_workers = min(get_analysis_batch_workers(), max(len(email_ids), 1))
        with ThreadPoolExecutor(max_workers=max_workers, thread_name_prefix="email-analysis") as executor:
            futures = [
                executor.submit(
                    _run_batch_email_job,
                    job_id=job_id,
                    email_id=email_id,
                    selected_subcriteria=list(selected_subcriteria),
                    force=force,
                )
                for email_id in email_ids
            ]
            for future in as_completed(futures):
                result = future.result()
                completed_count += 1
                job = db.get(Job, UUID(job_id))
                if job is None:
                    return
                job.progress_current = completed_count
                job.progress_total = len(email_ids)
                if result.get("status") == "failed_internal":
                    failed_internal_messages.append(str(result.get("error_message") or "Error interno"))
                db.commit()

        if job.batch_id:
            batch_emails = _load_batch_emails_for_mcdm(db, job.batch_id)
            _recompute_mcdm_for_emails(
                batch_emails,
                db=db,
                job=job,
                operation="analysis_batch_mcdm_refresh",
            )
            db.commit()

        if failed_internal_messages:
            job.status = "failed"
            job.error_message = failed_internal_messages[0]
        else:
            job.status = "completed"
            job.error_message = None
        job.finished_at = datetime.utcnow()
        db.commit()
    except Exception as exc:
        db.rollback()
        job = db.get(Job, UUID(job_id))
        if job is not None:
            job.status = "failed"
            job.error_message = str(exc)
            job.finished_at = datetime.utcnow()
            db.commit()
    finally:
        db.close()


def create_reanalyze_job(
    db: Session,
    background_tasks: BackgroundTasks,
    email: Email,
    subcriterion_key: str,
) -> Dict[str, Any]:
    if _email_is_discarded(email):
        raise ValueError("No se puede reanalizar un correo cancelado o descartado por error previo.")
    latest_job = _latest_job_for_email(db, email)
    if _job_is_active(latest_job):
        raise ValueError("No se puede lanzar otro analisis mientras ya hay un job activo para este correo.")

    definition = get_subcriterion_definition(subcriterion_key)
    selected = set(_sanitize_selected_subcriteria(email.selected_subcriteria))
    selected.add(definition["key"])
    email.selected_subcriteria = list(sorted(selected))
    job = Job(
        job_type="reanalyze_subcriterion",
        target_type="email",
        target_id=str(email.id),
        email_id=email.id,
        selected_subcriteria=[definition["key"]],
        status="queued",
        progress_current=0,
        progress_total=1,
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    background_tasks.add_task(run_email_job, str(job.id), str(email.id), [definition["key"]], True)
    return {"job_id": str(job.id), "subcriterion_key": definition["key"]}


def create_retry_email_job(
    db: Session,
    background_tasks: BackgroundTasks,
    email: Email,
) -> Dict[str, Any]:
    if not _email_is_discarded(email):
        raise ValueError("Solo se pueden reintentar correos cancelados o descartados por error previo.")
    latest_job = _latest_job_for_email(db, email)
    if _job_is_active(latest_job):
        raise ValueError("No se puede lanzar otro analisis mientras ya hay un job activo para este correo.")

    retry_subcriteria = _effective_selected_subcriteria(email)
    if not retry_subcriteria:
        raise ValueError("El correo no tiene subcriterios previos para reintentar.")

    _reset_email_processing_state(email)
    email.selected_subcriteria = _merge_selected_subcriteria(email.selected_subcriteria, retry_subcriteria)
    job = Job(
        job_type="retry_email",
        target_type="email",
        target_id=str(email.id),
        email_id=email.id,
        selected_subcriteria=retry_subcriteria,
        status="queued",
        progress_current=0,
        progress_total=len(retry_subcriteria),
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    background_tasks.add_task(run_email_job, str(job.id), str(email.id), retry_subcriteria, False)
    return {
        "job_id": str(job.id),
        "selected_subcriteria": retry_subcriteria,
    }


def create_analyze_missing_email_job(
    db: Session,
    background_tasks: BackgroundTasks,
    email: Email,
) -> Dict[str, Any]:
    if _email_is_discarded(email):
        raise ValueError("No se puede analizar un correo cancelado o descartado por error previo.")
    latest_job = _latest_job_for_email(db, email)
    if _job_is_active(latest_job):
        raise ValueError("No se puede lanzar otro analisis mientras ya hay un job activo para este correo.")

    missing = _missing_subcriteria_for_email(email)
    if not missing:
        raise ValueError("No hay subcriterios pendientes en este correo.")

    email.selected_subcriteria = _merge_selected_subcriteria(email.selected_subcriteria, missing)
    job = Job(
        job_type="analyze_missing_subcriteria",
        target_type="email",
        target_id=str(email.id),
        email_id=email.id,
        selected_subcriteria=missing,
        status="queued",
        progress_current=0,
        progress_total=len(missing),
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    background_tasks.add_task(run_email_job, str(job.id), str(email.id), missing, False)
    return {
        "job_id": str(job.id),
        "selected_subcriteria": missing,
    }


def create_analyze_missing_batch_job(
    db: Session,
    background_tasks: BackgroundTasks,
    batch: Batch,
) -> Dict[str, Any]:
    latest_job = _latest_job_for_batch(db, batch.id)
    if _job_is_active(latest_job):
        raise ValueError("No se puede lanzar otro analisis mientras ya hay un job activo en el lote.")
    processable_emails = _processable_batch_emails(batch)
    email_jobs = [job for email in processable_emails for job in email.jobs]
    if any(_job_is_active(job) for job in email_jobs):
        raise ValueError("No se puede lanzar otro analisis mientras hay jobs activos en correos del lote.")

    missing = _missing_subcriteria_for_batch(batch)
    if not missing:
        raise ValueError("No hay subcriterios pendientes en este lote.")

    email_ids: List[str] = []
    for email in processable_emails:
        email.selected_subcriteria = _merge_selected_subcriteria(email.selected_subcriteria, missing)
        email_ids.append(str(email.id))

    job = Job(
        job_type="analyze_missing_subcriteria",
        target_type="batch",
        target_id=str(batch.id),
        batch_id=batch.id,
        selected_subcriteria=missing,
        status="queued",
        progress_current=0,
        progress_total=len(email_ids),
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    background_tasks.add_task(run_batch_job, str(job.id), email_ids, missing, False)
    return {
        "job_id": str(job.id),
        "selected_subcriteria": missing,
    }


def run_retry_cancelled_batch_job(job_id: str, batch_id: str) -> None:
    session_factory = get_session_local()
    db = session_factory()
    try:
        job = db.get(Job, UUID(job_id))
        batch = db.get(Batch, UUID(batch_id))
        if job is None or batch is None:
            return

        job.status = "running"
        job.started_at = datetime.utcnow()
        db.commit()

        retryable_emails = _retryable_batch_emails(batch)
        job.progress_total = len(retryable_emails)
        job.progress_current = 0
        db.commit()

        retry_tasks: List[Tuple[str, List[str]]] = []
        for index, retryable_email in enumerate(retryable_emails, start=1):
            email = db.get(Email, retryable_email.id)
            if email is None:
                job.progress_current = index
                db.commit()
                continue

            retry_subcriteria = _effective_selected_subcriteria(email)
            if not retry_subcriteria:
                job.progress_current = index
                db.commit()
                continue

            _reset_email_processing_state(email)
            email.selected_subcriteria = _merge_selected_subcriteria(email.selected_subcriteria, retry_subcriteria)
            retry_tasks.append((str(email.id), list(retry_subcriteria)))

        db.commit()

        completed_count = job.progress_current
        failed_internal_messages: List[str] = []
        max_workers = min(get_analysis_batch_workers(), max(len(retry_tasks), 1))
        with ThreadPoolExecutor(max_workers=max_workers, thread_name_prefix="email-retry") as executor:
            futures = [
                executor.submit(
                    _run_batch_email_job,
                    job_id=job_id,
                    email_id=email_id,
                    selected_subcriteria=retry_subcriteria,
                    force=True,
                )
                for email_id, retry_subcriteria in retry_tasks
            ]
            for future in as_completed(futures):
                result = future.result()
                completed_count += 1
                job = db.get(Job, UUID(job_id))
                if job is None:
                    return
                job.progress_current = completed_count
                job.progress_total = len(retryable_emails)
                if result.get("status") == "failed_internal":
                    failed_internal_messages.append(str(result.get("error_message") or "Error interno"))
                db.commit()

        batch_emails = _load_batch_emails_for_mcdm(db, batch.id)
        _recompute_mcdm_for_emails(
            batch_emails,
            db=db,
            job=job,
            operation="retry_cancelled_batch_mcdm_refresh",
        )
        db.commit()

        if failed_internal_messages:
            job.status = "failed"
            job.error_message = failed_internal_messages[0]
        else:
            job.status = "completed"
            job.error_message = None
        job.finished_at = datetime.utcnow()
        db.commit()
    except Exception as exc:
        db.rollback()
        job = db.get(Job, UUID(job_id))
        if job is not None:
            job.status = "failed"
            job.error_message = str(exc)
            job.finished_at = datetime.utcnow()
            db.commit()
    finally:
        db.close()


def create_retry_cancelled_batch_job(
    db: Session,
    background_tasks: BackgroundTasks,
    batch: Batch,
) -> Dict[str, Any]:
    latest_job = _latest_job_for_batch(db, batch.id)
    if _job_is_active(latest_job):
        raise ValueError("No se puede lanzar otro analisis mientras ya hay un job activo en el lote.")

    email_jobs = [job for email in batch.emails for job in email.jobs]
    if any(_job_is_active(job) for job in email_jobs):
        raise ValueError("No se puede lanzar otro analisis mientras hay jobs activos en correos del lote.")

    retryable_emails = _retryable_batch_emails(batch)
    if not retryable_emails:
        raise ValueError("No hay correos cancelados para reintentar en este lote.")

    job = Job(
        job_type="retry_cancelled_batch",
        target_type="batch",
        target_id=str(batch.id),
        batch_id=batch.id,
        selected_subcriteria=[],
        status="queued",
        progress_current=0,
        progress_total=len(retryable_emails),
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    background_tasks.add_task(run_retry_cancelled_batch_job, str(job.id), str(batch.id))
    return {
        "job_id": str(job.id),
        "emails_total": len(retryable_emails),
    }


def run_batch_mcdm_recompute_job(job_id: str, batch_id: str) -> None:
    session_factory = get_session_local()
    db = session_factory()
    mcdm_started_at: Optional[datetime] = None
    mcdm_started_perf: Optional[float] = None
    mcdm_email_count = 0
    try:
        job = db.get(Job, UUID(job_id))
        batch_uuid = UUID(batch_id)
        if job is None:
            return

        job.status = "running"
        job.started_at = datetime.utcnow()
        db.commit()

        batch_emails = _load_batch_emails_for_mcdm(db, batch_uuid)
        job.progress_total = len(batch_emails)
        job.progress_current = 0
        mcdm_email_count = len(batch_emails)
        db.commit()

        mcdm_started_at = datetime.utcnow()
        mcdm_started_perf = time.perf_counter()
        processed_payloads: List[Dict[str, Any]] = []
        for email in batch_emails:
            raw_json = email.raw_json if isinstance(email.raw_json, dict) else {}
            numeric_values = raw_json.get("numeric_values")
            payload = dict(raw_json)
            if isinstance(numeric_values, dict):
                payload["numeric_values"] = dict(numeric_values)
            processed_payloads.append(payload)

        mcdm_blocks = compute_mcdm_blocks_for_emails(processed_payloads)
        for index, (email, processed_email, mcdm_block) in enumerate(
            zip(batch_emails, processed_payloads, mcdm_blocks),
            start=1,
        ):
            _apply_mcdm_block(email, processed_email, mcdm_block)
            job.progress_current = index
            db.commit()

        _persist_mcdm_runtime_metric(
            db,
            job=job,
            batch_id=batch_uuid,
            operation="recompute_batch_mcdm",
            started_at=mcdm_started_at,
            duration_ms=(time.perf_counter() - mcdm_started_perf) * 1000,
            email_count=mcdm_email_count,
        )
        job.status = "completed"
        job.finished_at = datetime.utcnow()
        job.error_message = None
        db.commit()
    except Exception as exc:
        db.rollback()
        job = db.get(Job, UUID(job_id))
        if job is not None:
            if mcdm_started_at is not None and mcdm_started_perf is not None:
                _persist_mcdm_runtime_metric(
                    db,
                    job=job,
                    batch_id=UUID(batch_id),
                    operation="recompute_batch_mcdm",
                    started_at=mcdm_started_at,
                    duration_ms=(time.perf_counter() - mcdm_started_perf) * 1000,
                    email_count=mcdm_email_count,
                    status="failed",
                    error_message=str(exc),
                )
            job.status = "failed"
            job.error_message = str(exc)
            job.finished_at = datetime.utcnow()
            db.commit()
    finally:
        db.close()


def create_recompute_batch_mcdm_job(
    db: Session,
    background_tasks: BackgroundTasks,
    batch: Batch,
) -> Dict[str, Any]:
    latest_job = _latest_job_for_batch(db, batch.id)
    if _job_is_active(latest_job):
        raise ValueError("No se puede recalcular el MCDM mientras hay un job activo en el lote.")
    processable_emails = _processable_batch_emails(batch)
    email_jobs = [job for email in processable_emails for job in email.jobs]
    if any(_job_is_active(job) for job in email_jobs):
        raise ValueError("No se puede recalcular el MCDM mientras hay jobs activos en correos del lote.")

    email_count = len(processable_emails)
    if email_count == 0:
        raise ValueError("No hay correos procesables en este lote.")
    job = Job(
        job_type="recompute_batch_mcdm",
        target_type="batch",
        target_id=str(batch.id),
        batch_id=batch.id,
        selected_subcriteria=[],
        status="queued",
        progress_current=0,
        progress_total=email_count,
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    background_tasks.add_task(run_batch_mcdm_recompute_job, str(job.id), str(batch.id))
    return {
        "job_id": str(job.id),
        "emails_total": email_count,
    }


def export_batch_mcdm_workbook(db: Session, batch: Batch) -> tuple[str, bytes]:
    from openpyxl import Workbook

    latest_job = _latest_job_for_batch(db, batch.id)
    if _job_is_active(latest_job) or _batch_has_active_email_jobs(batch):
        raise ValueError("No se puede exportar el Excel mientras hay jobs activos en el lote.")

    definitions = available_subcriteria()

    workbook = Workbook()
    workbook.remove(workbook.active)
    _append_export_sheet(
        workbook,
        "mcdm_scores_0_1",
        batch,
        definitions,
        db,
        score_mode=True,
    )
    _append_export_sheet(
        workbook,
        "numeric_values",
        batch,
        definitions,
        db,
        score_mode=False,
    )
    _append_weights_and_ranges_export_sheet(workbook, SUBCRITERIA_DEFINITIONS)
    analysis_metrics = _load_export_analysis_metrics(db, batch)
    api_call_metrics = _load_export_api_call_metrics(db, analysis_metrics)
    mcdm_runtime_metrics = _load_export_mcdm_runtime_metrics(db, batch)
    _append_timing_export_sheet(workbook, batch, analysis_metrics)
    _append_mcdm_runtime_export_sheet(workbook, mcdm_runtime_metrics)
    _append_api_call_export_sheet(workbook, batch, analysis_metrics, api_call_metrics)

    buffer = BytesIO()
    workbook.save(buffer)
    filename = f'{_safe_filename(batch.name)}_mcdm_export.xlsx'
    return filename, buffer.getvalue()


def _copy_email_analysis(source: Email, target: Email) -> None:
    source_enrichment = source.enrichment
    if source_enrichment is None:
        _create_empty_related(target)
        return

    target_enrichment = EmailEnrichment(
        mcdm_score=source_enrichment.mcdm_score,
        mcdm_is_mock=source_enrichment.mcdm_is_mock,
        mcdm_method=source_enrichment.mcdm_method,
    )
    for column in ENRICHMENT_JSON_COLUMNS.values():
        setattr(target_enrichment, column, deepcopy(getattr(source_enrichment, column)))

    source_vector = source_enrichment.vector
    target_vector = EmailMcdmVector()
    if source_vector is not None:
        target_vector.vector_version = source_vector.vector_version
        for field in VECTOR_FIELD_LABELS:
            setattr(target_vector, field, getattr(source_vector, field))

    target_enrichment.vector = target_vector
    target.enrichment = target_enrichment


def create_merged_batch(
    db: Session,
    *,
    batch_name: str,
    batch_ids: Sequence[UUID],
) -> Dict[str, Any]:
    batch_name = batch_name.strip()
    if not batch_name:
        raise ValueError("Debes indicar un nombre para el lote unido.")

    ordered_ids: List[UUID] = []
    seen_ids: set[UUID] = set()
    for batch_id in batch_ids:
        if batch_id in seen_ids:
            continue
        seen_ids.add(batch_id)
        ordered_ids.append(batch_id)

    if len(ordered_ids) < 2:
        raise ValueError("Selecciona al menos dos lotes para unir.")

    batches = (
        db.scalars(
            get_batch_query().where(Batch.id.in_(ordered_ids))
        )
        .unique()
        .all()
    )
    by_id = {batch.id: batch for batch in batches}
    missing = [str(batch_id) for batch_id in ordered_ids if batch_id not in by_id]
    if missing:
        raise ValueError(f"No se encontraron todos los lotes seleccionados: {', '.join(missing)}")

    ordered_batches = [by_id[batch_id] for batch_id in ordered_ids]
    active_jobs = [
        job
        for batch in ordered_batches
        for job in batch.jobs
        if _job_is_active(job)
    ]
    active_jobs.extend(
        job
        for batch in ordered_batches
        for email in batch.emails
        for job in email.jobs
        if _job_is_active(job)
    )
    if active_jobs:
        raise ValueError("No se pueden unir lotes mientras alguno tiene jobs activos.")

    latest_by_identity: Dict[str, Email] = {}
    for batch in ordered_batches:
        for email in batch.emails:
            identity = _email_merge_identity(email)
            current = latest_by_identity.get(identity)
            if current is None or _email_evaluation_sort_key(email) > _email_evaluation_sort_key(current):
                latest_by_identity[identity] = email

    source_emails = sorted(
        latest_by_identity.values(),
        key=lambda email: (_datetime_sort_value(email.created_at), str(email.id)),
    )
    if not source_emails:
        raise ValueError("Los lotes seleccionados no contienen correos.")

    merged_batch = Batch(name=batch_name, total_emails=len(source_emails))
    db.add(merged_batch)
    db.flush()

    email_ids: List[str] = []
    for source in source_emails:
        merged_email = Email(
            name=source.name,
            subject=source.subject,
            raw_json=deepcopy(source.raw_json),
            batch_id=merged_batch.id,
            selected_subcriteria=_effective_selected_subcriteria(source),
            processing_state=source.processing_state,
            processing_error_code=source.processing_error_code,
            processing_error_message=source.processing_error_message,
            processing_error_at=source.processing_error_at,
        )
        _copy_email_analysis(source, merged_email)
        db.add(merged_email)
        db.flush()
        email_ids.append(str(merged_email.id))

    _recompute_mcdm_for_emails(
        _load_batch_emails_for_mcdm(db, merged_batch.id),
        db=db,
        operation="merge_batch_mcdm_refresh",
    )
    db.commit()
    db.refresh(merged_batch)
    return {
        "batch_id": str(merged_batch.id),
        "email_ids": email_ids,
        "total_emails": len(email_ids),
        "deduplicated_emails": sum(len(batch.emails) for batch in ordered_batches) - len(email_ids),
    }


def delete_email_record(db: Session, email: Email) -> Dict[str, Any]:
    latest_job = _latest_job_for_email(db, email)
    if _job_is_active(latest_job):
        raise ValueError("No se puede eliminar un correo con un job en ejecucion.")

    analysis_metrics = db.scalars(
        select(EmailAnalysisMetric).where(EmailAnalysisMetric.email_id == email.id)
    ).all()
    for metric in analysis_metrics:
        db.delete(metric)

    jobs = db.scalars(select(Job).where(Job.email_id == email.id)).all()
    for job in jobs:
        db.delete(job)

    batch = email.batch
    if batch is not None:
        batch.total_emails = max(batch.total_emails - 1, 0)

    payload = {
        "email_id": str(email.id),
        "batch_id": str(batch.id) if batch is not None else None,
    }
    db.delete(email)
    db.commit()
    return payload


def delete_batch_record(db: Session, batch: Batch) -> Dict[str, Any]:
    batch_jobs = list(batch.jobs)
    email_jobs = [job for email in batch.emails for job in email.jobs]
    if any(_job_is_active(job) for job in [*batch_jobs, *email_jobs]):
        raise ValueError("No se puede eliminar un lote con un job en ejecucion.")

    email_ids = [email.id for email in batch.emails]
    if email_ids:
        analysis_metrics = db.scalars(
            select(EmailAnalysisMetric).where(
                or_(
                    EmailAnalysisMetric.batch_id == batch.id,
                    EmailAnalysisMetric.email_id.in_(email_ids),
                )
            )
        ).all()
        for metric in analysis_metrics:
            db.delete(metric)

    mcdm_metrics = db.scalars(
        select(McdmRuntimeMetric).where(McdmRuntimeMetric.batch_id == batch.id)
    ).all()
    for metric in mcdm_metrics:
        db.delete(metric)

    for job in batch_jobs:
        db.delete(job)

    if email_ids:
        jobs = db.scalars(select(Job).where(Job.email_id.in_(email_ids))).all()
        for job in jobs:
            db.delete(job)

    payload = {
        "batch_id": str(batch.id),
        "emails_deleted": len(email_ids),
    }
    db.delete(batch)
    db.commit()
    return payload


def get_email_summary_query() -> Select[tuple[Email]]:
    return select(Email).options(
        load_only(
            Email.id,
            Email.name,
            Email.subject,
            Email.created_at,
            Email.batch_id,
            Email.selected_subcriteria,
            Email.processing_state,
            Email.processing_error_code,
            Email.processing_error_message,
            Email.processing_error_at,
        ),
        selectinload(Email.batch).options(
            load_only(Batch.id, Batch.name),
            selectinload(Batch.jobs).load_only(*JOB_LOAD_ONLY_COLUMNS),
        ),
        selectinload(Email.jobs).load_only(*JOB_LOAD_ONLY_COLUMNS),
        selectinload(Email.enrichment).selectinload(EmailEnrichment.vector),
    )


def get_email_detail_query() -> Select[tuple[Email]]:
    return select(Email).options(
        load_only(
            Email.id,
            Email.name,
            Email.subject,
            Email.created_at,
            Email.batch_id,
            Email.selected_subcriteria,
            Email.raw_json,
            Email.processing_state,
            Email.processing_error_code,
            Email.processing_error_message,
            Email.processing_error_at,
        ),
        selectinload(Email.batch).options(
            load_only(Batch.id, Batch.name),
            selectinload(Batch.jobs).load_only(*JOB_LOAD_ONLY_COLUMNS),
        ),
        selectinload(Email.jobs).load_only(*JOB_LOAD_ONLY_COLUMNS),
        selectinload(Email.enrichment).selectinload(EmailEnrichment.vector),
    )


def get_batch_query() -> Select[tuple[Batch]]:
    return select(Batch).options(
        load_only(Batch.id, Batch.name, Batch.created_at, Batch.total_emails),
        selectinload(Batch.jobs).load_only(*JOB_LOAD_ONLY_COLUMNS),
        selectinload(Batch.emails).options(
            load_only(
                Email.id,
                Email.name,
                Email.subject,
                Email.created_at,
                Email.batch_id,
                Email.selected_subcriteria,
                Email.processing_state,
                Email.processing_error_code,
                Email.processing_error_message,
                Email.processing_error_at,
            ),
            selectinload(Email.batch).load_only(Batch.id, Batch.name),
            selectinload(Email.jobs).load_only(*JOB_LOAD_ONLY_COLUMNS),
            selectinload(Email.enrichment).selectinload(EmailEnrichment.vector),
        ),
    )


def get_batch_summary_query() -> Select[tuple[Batch]]:
    return select(Batch).options(
        load_only(Batch.id, Batch.name, Batch.created_at, Batch.total_emails),
        selectinload(Batch.jobs).load_only(*JOB_LOAD_ONLY_COLUMNS),
        selectinload(Batch.emails).options(
            load_only(
                Email.id,
                Email.selected_subcriteria,
                Email.processing_state,
            ),
            selectinload(Email.jobs).load_only(*JOB_LOAD_ONLY_COLUMNS),
            selectinload(Email.enrichment),
        ),
    )


__all__ = [
    "available_subcriteria",
    "read_uploads",
    "create_single_upload",
    "create_batch_upload",
    "create_batch_upload_from_files",
    "create_pending_batch_upload",
    "append_batch_upload_files",
    "finalize_batch_upload",
    "create_reanalyze_job",
    "create_retry_email_job",
    "create_analyze_missing_email_job",
    "create_analyze_missing_batch_job",
    "create_retry_cancelled_batch_job",
    "create_recompute_batch_mcdm_job",
    "create_merged_batch",
    "delete_email_record",
    "delete_batch_record",
    "serialize_email_summary",
    "serialize_email_detail",
    "serialize_batch_summary",
    "serialize_batch_detail",
    "export_batch_mcdm_workbook",
    "batch_score_analytics_overview",
    "batch_subcriterion_analytics",
    "get_email_summary_query",
    "get_email_detail_query",
    "get_batch_summary_query",
    "get_batch_query",
    "_serialize_job",
]
